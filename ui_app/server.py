import asyncio
import csv
import json
import getpass
import io
import os
import re
import shutil
import sys
import uuid
import zipfile
import urllib.error
import urllib.request
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from fastapi import Body, FastAPI, File, Form, HTTPException, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse

try:
    from openpyxl import Workbook, load_workbook
except Exception:
    Workbook = None
    load_workbook = None


ROOT = Path(__file__).resolve().parents[1]
SCRIPT_PATH = ROOT / "precision-auto-playwright-batch.py"
UPLOAD_DIR = ROOT / "ui_uploads"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
HISTORY_DIR = UPLOAD_DIR / "task_history"
HISTORY_DIR.mkdir(parents=True, exist_ok=True)
DEFAULT_DATA_CSV = ROOT / "data" / "plans.csv"
# 网页“Excel模板”下载优先使用用户指定模板文件
CUSTOM_EXPORT_TEMPLATE_XLSX = Path("/Users/liminrong/Downloads/精准营销任务模板（导入模板）.xlsx")


HEADER_EN_TO_CN: Dict[str, str] = {
    "name": "计划名称",
    "region": "计划区域",
    "theme": "营销主题",
    "scene_type": "场景类型",
    "plan_type": "计划类型",
    "push_content": "推送内容",
    "start_time": "计划开始时间",
    "end_time": "计划结束时间",
    "trigger_type": "触发方式",
    "send_time": "发送时间",
    "global_limit": "全局触达限制",
    "create_url": "创建链接",
    "group_name": "分群名称",
    "main_operating_area": "主消费营运区",
    "main_store_file_path": "主消费门店文件路径",
    "step2_store_file_path": "第2步门店信息文件路径",
    "step2_product_file_path": "第2步商品编码文件路径",
    "coupon_ids": "券规则ID",
    "coupon_ids_sheet_ref": "已领或已使用券规则ID",
    "purchase_target_product_code": "购买目标商品编码",
    "sms_content": "短信内容",
    "step3_end_time": "员工任务结束时间",
    "distribution_mode": "社群任务分配方式",
    "executor_employees": "执行员工",
    "send_content": "发送内容",
    "group_send_name": "下发群名",
    "channels": "发送渠道",
    "plan_image_id": "计划图片ID",
    "moments_add_images": "朋友圈是否上传图片",
    "moments_image_paths": "朋友圈图片路径(用|分隔)",
    "upload_stores": "是否上传门店",
    "store_file_path": "门店文件路径",
    "msg_add_mini_program": "会员通消息是否添加小程序",
    "msg_mini_program_name": "1对1-小程序名称",
    "msg_mini_program_title": "1对1-小程序标题",
    "msg_mini_program_cover_path": "小程序封面路径",
    "msg_mini_program_page_path": "1对1-小程序链接",
}
HEADER_CN_TO_EN: Dict[str, str] = {v: k for k, v in HEADER_EN_TO_CN.items()}
HEADER_CN_TO_EN.update({
    "第3步结束时间": "step3_end_time",
    "分配方式": "distribution_mode",
    "发送内容": "send_content",
    "短信内容": "sms_content",
    "主消费运营区": "main_operating_area",
    "计划图片id": "plan_image_id",
    "计划图片Id": "plan_image_id",
    "1对1-小程序功能页面": "msg_mini_program_page_path",
})

CHANNEL_CODE_TO_NAME: Dict[str, str] = {
    "1": "短信",
    "2": "会员通-发客户消息",
    "3": "会员通-发客户朋友圈",
    "4": "会员通-发送社群",
}
CHANNEL_ALIAS_TO_NAME: Dict[str, str] = {
    "短信": "短信",
    "会员通-发客户消息": "会员通-发客户消息",
    "会员通发客户消息": "会员通-发客户消息",
    "会员通-发客户朋友圈": "会员通-发客户朋友圈",
    "会员通客户朋友圈": "会员通-发客户朋友圈",
    "会员通发客户朋友圈": "会员通-发客户朋友圈",
    "会员通-发送社群": "会员通-发送社群",
    "会员通发送社群": "会员通-发送社群",
}
TEMPLATE_HIDE_FIELDS = {
    "group_name",
    "moments_add_images",
    "moments_image_paths",
    "upload_stores",
    "store_file_path",
    "main_store_file_path",
    "step2_store_file_path",
    "step2_product_file_path",
    "msg_add_mini_program",
    "msg_mini_program_cover_path",
    "coupon_ids",
    "use_recommend",
    "set_target",
    "update_type",
    "sms_content",
    "send_content",
    "group_name",
    "region",
    "trigger_type",
    "group_send_name",
    "scene_type",
    "plan_type",
}

DEFAULT_CREATE_URL_BY_CHANNEL: Dict[str, str] = {
    "短信": "https://precision.dslyy.com/admin#/marketingTemplate/use?useId=599702746907561984",
    "会员通-发客户消息": "https://precision.dslyy.com/admin#/marketingTemplate/use?useId=594094287227023360",
    "会员通-发客户朋友圈": "https://precision.dslyy.com/admin#/marketingTemplate/use?useId=599702926159527936",
    "会员通-发送社群": "https://precision.dslyy.com/admin#/marketingPlan/addcommunityPlan?checkType=add",
}
DEFAULT_CREATE_URL_COMBO: Dict[str, str] = {
    "短信|会员通-发客户消息": "https://precision.dslyy.com/admin#/marketingTemplate/use?useId=600035736992907264",
}


def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def _task_history_path(task_id: str) -> Path:
    return HISTORY_DIR / f"{task_id}.json"


def _task_history_log_path(task_id: str) -> Path:
    return HISTORY_DIR / f"{task_id}.log"


def _task_failure_index_path() -> Path:
    return HISTORY_DIR / "failed_tasks.ndjson"


def _safe_tail(lines: List[str], size: int = 60) -> List[str]:
    if not lines:
        return []
    return lines[-size:]


def _extract_error_summary(task: "Task") -> str:
    # Prefer explicit "错误:" line from runtime logs.
    for ln in reversed(task.logs):
        s = (ln or "").strip()
        if "错误:" in s:
            return s
        if s.startswith("Error:") or s.startswith("Exception:"):
            return s
    return task.error or "任务失败，详情见日志"


def _persist_task_history(task: "Task") -> None:
    try:
        payload = task.to_dict()
        payload["logs"] = task.logs
        _task_history_path(task.id).write_text(
            json.dumps(payload, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        _task_history_log_path(task.id).write_text("\n".join(task.logs), encoding="utf-8")
    except Exception:
        pass


def _append_failure_index(task: "Task") -> None:
    try:
        rec = {
            "id": task.id,
            "filename": task.filename,
            "plan_name": task.plan_name_display,
            "channels": task.channel_display,
            "status": task.status,
            "error_summary": _extract_error_summary(task),
            "started_at": task.started_at,
            "ended_at": task.ended_at,
            "history_json": str(_task_history_path(task.id)),
            "history_log": str(_task_history_log_path(task.id)),
        }
        with _task_failure_index_path().open("a", encoding="utf-8") as f:
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")
    except Exception:
        pass


def parse_int(val: str, default: int = 0) -> int:
    try:
        return int(val)
    except Exception:
        return default


def _default_headers() -> List[str]:
    return [
        "name",
        "channels",
        "region",
        "theme",
        "scene_type",
        "plan_type",
        "push_content",
        "start_time",
        "end_time",
        "trigger_type",
        "send_time",
        "global_limit",
        "create_url",
        "main_operating_area",
        "main_store_file_path",
        "step2_store_file_path",
        "step2_product_file_path",
        "purchase_target_product_code",
        "coupon_ids",
        "coupon_ids_sheet_ref",
        "sms_content",
        "step3_end_time",
        "distribution_mode",
        "executor_employees",
        "group_send_name",
        "send_content",
        "plan_image_id",
        "moments_add_images",
        "moments_image_paths",
        "upload_stores",
        "store_file_path",
        "msg_add_mini_program",
        "msg_mini_program_name",
        "msg_mini_program_title",
        "msg_mini_program_cover_path",
        "msg_mini_program_page_path",
    ]


def load_template_headers_and_sample() -> tuple[List[str], List[str]]:
    # 模板字段固定使用系统标准字段，避免受本地旧CSV表头污染（导致下载模板出现英文/脏字段）
    headers = _default_headers()
    sample_map: Dict[str, str] = {}
    if DEFAULT_DATA_CSV.exists():
        for enc in ("utf-8-sig", "utf-8", "gbk"):
            try:
                with DEFAULT_DATA_CSV.open("r", encoding=enc, newline="") as f:
                    rows = list(csv.reader(f))
                if not rows:
                    continue
                raw_headers = [str(x or "").strip() for x in rows[0]]
                raw_sample = [str(x or "").strip() for x in (rows[1] if len(rows) > 1 else [""] * len(raw_headers))]
                norm_headers = [HEADER_CN_TO_EN.get(h, h) for h in raw_headers]
                sample_map = {
                    h: (raw_sample[idx] if idx < len(raw_sample) else "")
                    for idx, h in enumerate(norm_headers)
                }
                break
            except Exception:
                continue
    sample = [sample_map.get(h, "") for h in headers]
    return headers, sample


def _filter_template_fields(headers: List[str], sample: List[str]) -> tuple[List[str], List[str]]:
    keep = [idx for idx, h in enumerate(headers) if h not in TEMPLATE_HIDE_FIELDS]
    out_headers = [headers[idx] for idx in keep]
    out_sample = [sample[idx] if idx < len(sample) else "" for idx in keep]
    if "create_url" in out_headers:
        i = out_headers.index("create_url")
        h = out_headers.pop(i)
        v = out_sample.pop(i)
        out_headers.append(h)
        out_sample.append(v)
    # 业务引导示例：第1步营销主题支持多选填写
    if "theme" in out_headers:
        idx = out_headers.index("theme")
        out_sample[idx] = "其他、26年3月积分换券"
    if "channels" in out_headers:
        idx = out_headers.index("channels")
        out_sample[idx] = "会员通-发客户消息"
    if "create_url" in out_headers:
        idx = out_headers.index("create_url")
        out_sample[idx] = "https://precision.dslyy.com/admin#/marketingTemplate/use?useId=594094287227023360"
    if "push_content" in out_headers:
        idx = out_headers.index("push_content")
        out_sample[idx] = "示例推送内容（按发送渠道自动映射到短信内容或发送内容）"
    # 业务引导示例：第2步主消费营运区支持多区域填写
    if "main_operating_area" in out_headers:
        idx = out_headers.index("main_operating_area")
        out_sample[idx] = "辽宁省区、九江、南昌、广州二"
    if "executor_employees" in out_headers:
        idx = out_headers.index("executor_employees")
        out_sample[idx] = "《目标门店 1》"
    if "purchase_target_product_code" in out_headers:
        idx = out_headers.index("purchase_target_product_code")
        out_sample[idx] = "《目标商品 1》"
    if "coupon_ids_sheet_ref" in out_headers:
        idx = out_headers.index("coupon_ids_sheet_ref")
        out_sample[idx] = "《券规则 ID 1》"
    if "msg_mini_program_name" in out_headers:
        out_sample[out_headers.index("msg_mini_program_name")] = "大参林健康"
    if "start_time" in out_headers:
        out_sample[out_headers.index("start_time")] = "2026-04-01"
    if "end_time" in out_headers:
        out_sample[out_headers.index("end_time")] = "2026-04-10"
    if "step3_end_time" in out_headers:
        out_sample[out_headers.index("step3_end_time")] = "2026-04-10"
    return out_headers, out_sample


def write_template_csv(path: Path) -> None:
    headers, sample = load_template_headers_and_sample()
    headers, sample = _filter_template_fields(headers, sample)
    cn_headers = [HEADER_EN_TO_CN.get(h, h) for h in headers]
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(cn_headers)
        writer.writerow(sample)


def write_template_xlsx(path: Path) -> None:
    if Workbook is None:
        raise RuntimeError("openpyxl is not installed")
    headers, sample = load_template_headers_and_sample()
    headers, sample = _filter_template_fields(headers, sample)
    cn_headers = [HEADER_EN_TO_CN.get(h, h) for h in headers]
    wb = Workbook()
    ws = wb.active
    ws.title = "任务文件"
    ws.append(cn_headers)
    ws.append(sample)

    def _row_from_headers(hs: List[str], vals: Dict[str, str]) -> List[str]:
        return [vals.get(h, "") for h in hs]

    # 业务示例：用于给非技术同事直接参考填写（不参与程序逻辑判断）
    ws_example = wb.create_sheet("任务文件（示例）")
    ws_example.append(cn_headers)
    ws_example.append(_row_from_headers(cn_headers, {
        "计划名称": "测试2-企微1对1",
        "发送渠道": "会员通-发客户消息",
        "营销主题": "其他、26年3月积分换券",
        "推送内容": "测试2-企微1对1",
        "计划开始时间": "2026-04-01",
        "计划结束时间": "2026-04-10",
        "发送时间": "2026-04-08 08:00:00",
        "全局触达限制": "限制",
        "创建链接": "https://precision.dslyy.com/admin#/marketingTemplate/use?useId=594094287227023360",
        "主消费营运区": "《目标门店 1》",
        "购买目标商品编码": "《目标商品 1》",
        "已领或已使用券规则ID": "《券规则 ID 1》",
        "员工任务结束时间": "2026-04-10",
        "执行员工": "西北大区、湖北省区",
        "计划图片ID": "2",
        "1对1-小程序名称": "大参林健康",
        "1对1-小程序标题": "测试1-卡片",
        "1对1-小程序链接": "apps/member/integralMall/pages/home/index",
    }))
    ws_example.append(_row_from_headers(cn_headers, {
        "计划名称": "测试3-朋友圈",
        "发送渠道": "会员通-发客户朋友圈",
        "营销主题": "其他、26年3月积分换券",
        "推送内容": "测试3-朋友圈",
        "计划开始时间": "2026-04-01",
        "计划结束时间": "2026-04-10",
        "发送时间": "2026-04-08 08:00:00",
        "全局触达限制": "不限制",
        "创建链接": "https://precision.dslyy.com/admin#/marketingTemplate/use?useId=599702926159527936",
        "主消费营运区": "华南大区",
        "已领或已使用券规则ID": "《券规则 ID 1》",
        "员工任务结束时间": "2026-04-10",
        "执行员工": "黑龙江省区、武汉营运区",
        "计划图片ID": "3",
    }))
    ws_example.append(_row_from_headers(cn_headers, {
        "计划名称": "测试4-短信",
        "发送渠道": "短信",
        "营销主题": "其他、26年3月积分换券",
        "推送内容": "测试4-短信",
        "计划开始时间": "2026-04-01",
        "计划结束时间": "2026-04-10",
        "发送时间": "2026-04-08 08:00:00",
        "全局触达限制": "限制",
        "创建链接": "https://precision.dslyy.com/admin#/marketingTemplate/use?useId=599702746907561984",
        "主消费营运区": "来宾、华中大区",
        "购买目标商品编码": "《目标商品 1》",
        "已领或已使用券规则ID": "《券规则 ID 1》",
        "计划图片ID": "4",
    }))

    # 社群任务：独立sheet读取（发送渠道=会员通-发送社群）
    community_headers = [
        "计划名称", "发送渠道", "营销主题", "场景类型", "计划类型",
        "推送内容", "计划开始时间", "计划结束时间", "发送时间",
        "员工任务结束时间", "社群任务分配方式", "执行员工",
        "计划图片ID", "1对1-小程序名称", "1对1-小程序标题", "1对1-小程序链接",
    ]
    ws_community = wb.create_sheet("社群任务")
    ws_community.append(community_headers)
    ws_community.append([
        "测试1-社群", "会员通-发送社群", "其他、会员生日礼", "会员营销", "会员权益",
        "测试1-社群", "2026-04-01", "2026-04-10", "2026-04-08 08:00:00",
        "2026-04-10", "导入门店", "《目标门店 1》",
        "1", "大参林健康", "社群小程序示例", "pages/index/index",
    ])
    ws_store_1 = wb.create_sheet("目标门店 1")
    ws_store_1.append(["门店编码"])
    ws_store_1.append(["1001010022"])
    ws_store_1.append(["1001010026"])
    ws_product_1 = wb.create_sheet("目标商品 1")
    ws_product_1.append(["商品编码"])
    ws_product_1.append(["1010002"])
    ws_product_1.append(["1012058"])
    ws_coupon = wb.create_sheet("券规则 ID 1")
    ws_coupon.append(["券规则ID"])
    ws_coupon.append(["1-20000005313"])
    ws_coupon.append(["1-20000005475"])
    wb.save(path)
    wb.close()


def write_community_template_xlsx(path: Path) -> None:
    """社群专用模板：单Excel多sheet（任务文件/目标门店），仅保留社群最小必填字段。"""
    if Workbook is None:
        raise RuntimeError("openpyxl is not installed")
    wb = Workbook()
    ws = wb.active
    ws.title = "任务文件"
    headers = [
        "计划名称", "发送渠道", "营销主题", "场景类型", "计划类型",
        "计划开始时间", "计划结束时间", "发送时间", "员工任务结束时间",
        "社群任务分配方式", "执行员工", "推送内容",
        "1对1-小程序名称", "1对1-小程序标题", "1对1-小程序链接",
    ]
    ws.append(headers)
    ws.append([
        "专属社群测试模板（自动化）", "会员通-发送社群", "其他、会员生日礼", "会员营销", "会员权益",
        "2026-03-20", "2026-03-31", "2026-03-22 08:00:00",
        "2026-03-31", "按条件筛选客户群", "黑龙江省区、武汉营运区", "社群自动化测试内容",
        "大参林健康", "社群小程序示例", "pages/index/index",
    ])
    ws_store = wb.create_sheet("目标门店 1")
    ws_store.append(["门店编码"])
    ws_store.append(["2000081179"])
    ws_product = wb.create_sheet("目标商品 1")
    ws_product.append(["商品编码"])
    ws_product.append(["1012058"])
    wb.save(path)
    wb.close()


def normalize_uploaded_csv_headers(dst_csv: Path) -> None:
    """将上传文件中的中文表头标准化为脚本内部英文表头。"""
    try:
        with dst_csv.open("r", encoding="utf-8-sig", newline="") as f:
            rows = list(csv.reader(f))
        if not rows:
            return
        raw_headers = [str(x or "").strip() for x in rows[0]]
        normalized: List[str] = []
        for h in raw_headers:
            if h == "第3步渠道(可多选)":
                normalized.append("channels")
            else:
                normalized.append(HEADER_CN_TO_EN.get(h, h))
        if normalized == raw_headers:
            return
        rows[0] = normalized
        with dst_csv.open("w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerows(rows)
    except Exception:
        # 不阻断上传；后续由脚本校验字段
        return


def _normalize_channel_text(raw: str) -> str:
    parts = [p.strip() for p in re.split(r"[|,，、/]+", str(raw or "")) if p.strip()]
    out: List[str] = []
    for p in parts:
        v = CHANNEL_CODE_TO_NAME.get(p) or CHANNEL_ALIAS_TO_NAME.get(p) or p
        if v not in out:
            out.append(v)
    return "、".join(out)


def normalize_channels_in_csv(dst_csv: Path) -> None:
    """支持文件中发送渠道填写序号（1/2/3/4），统一标准化为中文渠道名称。"""
    try:
        with dst_csv.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
            headers = list(reader.fieldnames or [])
        if not headers or "channels" not in headers:
            return
        changed = False
        for row in rows:
            old = str(row.get("channels", "") or "")
            new = _normalize_channel_text(old)
            if new and new != old:
                row["channels"] = new
                changed = True
        if not changed:
            return
        with dst_csv.open("w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            for row in rows:
                writer.writerow({k: row.get(k, "") for k in headers})
    except Exception:
        return


def _parse_dt_for_upload(raw: str, *, end_of_day_for_date_only: bool = False) -> datetime:
    s = str(raw or "").strip()
    if not s:
        raise ValueError("空值")
    fmts = (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
        "%Y-%m-%d",
        "%Y/%m/%d",
    )
    for fmt in fmts:
        try:
            dt = datetime.strptime(s, fmt)
            if fmt in ("%Y-%m-%d", "%Y/%m/%d") and end_of_day_for_date_only:
                dt = dt.replace(hour=23, minute=59, second=59)
            return dt
        except Exception:
            continue
    raise ValueError(f"无法识别时间格式: {s}")


def prevalidate_csv_time_fields(dst_csv: Path) -> None:
    """
    上传阶段前置校验，避免任务入列后才失败：
    - 计划起止时间不能超过14天，且结束>=开始
    - 发送时间不能小于当前时间
    - 员工任务结束时间（若有）不能小于当前时间
    """
    with dst_csv.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        headers = list(reader.fieldnames or [])
    if not headers:
        return

    now_dt = datetime.now()
    for i, row in enumerate(rows, start=1):
        if not any(str(v or "").strip() for v in row.values()):
            continue

        start_s = str(row.get("start_time", "") or "").strip()
        end_s = str(row.get("end_time", "") or "").strip()
        send_s = str(row.get("send_time", "") or "").strip()
        step3_end_s = str(row.get("step3_end_time", "") or "").strip()

        if start_s and end_s:
            try:
                st = _parse_dt_for_upload(start_s)
                et = _parse_dt_for_upload(end_s, end_of_day_for_date_only=True)
                if et < st:
                    raise HTTPException(status_code=400, detail=f"第{i}行：计划结束时间早于开始时间")
                if (et - st).total_seconds() > 14 * 24 * 3600:
                    raise HTTPException(status_code=400, detail=f"第{i}行：计划时间起止不能超过14天")
            except HTTPException:
                raise
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"第{i}行：计划时间格式错误（计划开始时间/计划结束时间）: {e}")

        if send_s:
            try:
                send_dt = _parse_dt_for_upload(send_s)
                if send_dt < now_dt:
                    raise HTTPException(
                        status_code=400,
                        detail=(
                            f"第{i}行：发送时间不能小于当前时间"
                            f"（send_time={send_dt.strftime('%Y-%m-%d %H:%M:%S')}，"
                            f"now={now_dt.strftime('%Y-%m-%d %H:%M:%S')}）"
                        ),
                    )
            except HTTPException:
                raise
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"第{i}行：发送时间格式错误（发送时间）: {e}")

        if step3_end_s:
            try:
                step3_dt = _parse_dt_for_upload(step3_end_s, end_of_day_for_date_only=True)
                if step3_dt < now_dt:
                    raise HTTPException(
                        status_code=400,
                        detail=(
                            f"第{i}行：员工任务结束时间不能小于当前时间"
                            f"（step3_end_time={step3_dt.strftime('%Y-%m-%d %H:%M:%S')}，"
                            f"now={now_dt.strftime('%Y-%m-%d %H:%M:%S')}）"
                        ),
                    )
            except HTTPException:
                raise
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"第{i}行：员工任务结束时间格式错误: {e}")


def _norm_sheet_name(name: str) -> str:
    return re.sub(r"\s+", "", str(name or "")).strip().lower()


def _extract_book_title_ref(raw: str) -> str:
    m = re.search(r"《\s*([^》]+?)\s*》", str(raw or ""))
    return (m.group(1).strip() if m else "")


def _collect_sheet_first_col_values(values: List[List[str]]) -> List[str]:
    if not values:
        return []
    out: List[str] = []
    for i, row in enumerate(values):
        if not row:
            continue
        first = str(row[0] or "").strip()
        if not first:
            continue
        if i == 0 and ("券规则" in first or "id" in first.lower()):
            continue
        if first not in out:
            out.append(first)
    return out


def _norm_col_text(v: str) -> str:
    return re.sub(r"\s+", "", str(v or "")).strip().lower()


def _ensure_sheet_has_key_header(
    values: List[List[str]],
    aliases: List[str],
    row_no: int,
    sheet_title: str,
    field_desc: str,
) -> None:
    """
    书名号引用的 sheet 必须包含关键列名，否则直接阻断，避免后续误识别。
    """
    if not values or not values[0]:
        raise HTTPException(status_code=400, detail=f"第{row_no}行：sheet《{sheet_title}》为空，缺少{field_desc}关键列")
    header = [_norm_col_text(x) for x in values[0]]
    alias_norm = {_norm_col_text(a) for a in aliases if str(a or "").strip()}
    if not any(h in alias_norm for h in header):
        alias_text = " / ".join(aliases)
        raise HTTPException(
            status_code=400,
            detail=f"第{row_no}行：sheet《{sheet_title}》缺少关键列（{alias_text}）",
        )


def apply_unified_field_mapping_and_refs(
    dst_csv: Path,
    task_id: str,
    step3_channels: str,
    sheet_assets: Dict[str, dict],
) -> None:
    """
    统一模板字段映射与书名号sheet引用处理：
    - 推送内容 -> 按发送渠道映射到 sms_content / send_content
    - 执行员工《目标门店X》 -> 自动映射 store_file_path + upload_stores=是
    - 购买目标商品编码《目标商品X》 -> step2_product_file_path
    - 已领或已使用券规则ID《券规则IDX》 -> coupon_ids（用/拼接）
    - 创建链接规则：社群空值自动补默认；非社群空值报错阻断
    """
    with dst_csv.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        headers = list(reader.fieldnames or [])
    if not headers:
        return

    for col in (
        "sms_content",
        "send_content",
        "executor_include_franchise",
        "upload_stores",
        "store_file_path",
        "main_store_file_path",
        "step2_store_file_path",
        "step2_product_file_path",
        "coupon_ids",
        "channels",
        "create_url",
        "msg_add_mini_program",
        "msg_mini_program_cover_path",
    ):
        if col not in headers:
            headers.append(col)

    ui_channels = _normalize_channel_text(step3_channels or "")
    community_default = "https://precision.dslyy.com/admin#/marketingPlan/addcommunityPlan?checkType=add"

    def _default_create_url_for_parts(parts: List[str]) -> str:
        uniq = []
        for p in parts:
            if p not in uniq:
                uniq.append(p)
        combo_key = "|".join(sorted(uniq))
        if combo_key in DEFAULT_CREATE_URL_COMBO:
            return DEFAULT_CREATE_URL_COMBO[combo_key]
        for p in uniq:
            if p in DEFAULT_CREATE_URL_BY_CHANNEL:
                return DEFAULT_CREATE_URL_BY_CHANNEL[p]
        return ""

    def _pick_default_mini_cover() -> str:
        patterns = ("cover_*.JPG", "cover_*.jpg", "cover_*.jpeg", "cover_*.png")
        files: List[Path] = []
        for pat in patterns:
            files.extend(UPLOAD_DIR.glob(f"*_mini_program/{pat}"))
        if not files:
            return ""
        files = sorted(files, key=lambda x: x.stat().st_mtime, reverse=True)
        return str(files[0].resolve())

    def _save_sheet_blob_for_store(sheet_title: str, row_no: int) -> str:
        asset = sheet_assets.get(_norm_sheet_name(sheet_title))
        if not asset:
            raise HTTPException(status_code=400, detail=f"第{row_no}行：未找到sheet《{sheet_title}》")
        _ensure_sheet_has_key_header(
            asset.get("rows", []),
            aliases=["门店编码", "门店code", "storecode"],
            row_no=row_no,
            sheet_title=sheet_title,
            field_desc="目标门店",
        )
        return save_uploaded_store_file(f"{task_id}_r{row_no}", (asset["filename"], asset["bytes"]))

    def _save_sheet_blob_for_step2_product(sheet_title: str, row_no: int) -> str:
        asset = sheet_assets.get(_norm_sheet_name(sheet_title))
        if not asset:
            raise HTTPException(status_code=400, detail=f"第{row_no}行：未找到sheet《{sheet_title}》")
        _ensure_sheet_has_key_header(
            asset.get("rows", []),
            aliases=["商品编码", "商品code", "productcode"],
            row_no=row_no,
            sheet_title=sheet_title,
            field_desc="目标商品",
        )
        return save_uploaded_step2_product_file(f"{task_id}_r{row_no}", (asset["filename"], asset["bytes"]))

    normalized_rows: List[Dict[str, str]] = []
    for row in rows:
        # 跳过 Excel 末尾空行（避免误报“发送渠道不能为空”）
        if not any(str(v or "").strip() for v in row.values()):
            continue
        normalized_rows.append(row)

    for idx, row in enumerate(normalized_rows, 1):
        row_channels = _normalize_channel_text(str(row.get("channels", "") or "").strip())
        channel_scope = row_channels or ui_channels
        parts = [p.strip() for p in re.split(r"[|,，、/]+", channel_scope) if p.strip()]
        is_community = "会员通-发送社群" in parts
        is_moments = "会员通-发客户朋友圈" in parts
        is_customer_msg = "会员通-发客户消息" in parts

        # 默认值补齐（模板已简化，避免缺列后脚本失败）
        if not str(row.get("region", "") or "").strip():
            row["region"] = "省区"
        if not str(row.get("trigger_type", "") or "").strip():
            row["trigger_type"] = "定时-单次任务"
        if not str(row.get("group_send_name", "") or "").strip():
            row["group_send_name"] = "福利"
        if is_community or is_moments:
            row["global_limit"] = "不限制"

        # 日期字段自动补时分秒
        def _norm_dt_text(raw: str, *, end_of_day_for_date_only: bool) -> str:
            s = str(raw or "").strip()
            if not s:
                return ""
            dt = _parse_dt_for_upload(s, end_of_day_for_date_only=end_of_day_for_date_only)
            return dt.strftime("%Y-%m-%d %H:%M:%S")

        for key, eod in (
            ("start_time", False),
            ("end_time", True),
            ("step3_end_time", True),
        ):
            val = str(row.get(key, "") or "").strip()
            if val:
                try:
                    row[key] = _norm_dt_text(val, end_of_day_for_date_only=eod)
                except Exception as e:
                    label = HEADER_EN_TO_CN.get(key, key)
                    raise HTTPException(status_code=400, detail=f"第{idx}行：{label}格式错误: {e}")

        # 发送渠道严格校验
        if not parts:
            raise HTTPException(status_code=400, detail=f"第{idx}行：发送渠道不能为空")
        row["channels"] = "、".join(parts)

        # 社群任务：场景类型、计划类型必填（仅社群要求）
        if is_community:
            if not str(row.get("scene_type", "") or "").strip():
                raise HTTPException(status_code=400, detail=f"第{idx}行：社群任务“场景类型”不能为空")
            if not str(row.get("plan_type", "") or "").strip():
                raise HTTPException(status_code=400, detail=f"第{idx}行：社群任务“计划类型”不能为空")

        # 创建链接规则：任务文件中不允许手填，统一按渠道自动赋值
        auto_url = _default_create_url_for_parts(parts) or community_default
        # 强制覆盖，避免误链路
        row["create_url"] = auto_url

        # 推送内容路由
        push_content = str(row.get("push_content", "") or "").strip()
        if push_content:
            if "短信" in parts:
                row["sms_content"] = push_content
            if any(p in {"会员通-发客户消息", "会员通-发客户朋友圈", "会员通-发送社群"} for p in parts):
                row["send_content"] = push_content

        # 主消费营运区书名号引用 -> 第2步门店信息sheet（兼容用户把《目标门店X》写在该字段）
        main_area_ref = _extract_book_title_ref(row.get("main_operating_area", ""))
        if main_area_ref:
            step2_store_path = _save_sheet_blob_for_store(main_area_ref, idx)
            row["main_store_file_path"] = step2_store_path
            row["step2_store_file_path"] = step2_store_path
            # 书名号引用场景下不再走“主消费营运区树节点选择”
            row["main_operating_area"] = ""
            # 会员通1对1/朋友圈：主消费营运区《xxx》同步作为执行员工“上传门店”来源
            if is_customer_msg or is_moments:
                row["upload_stores"] = "是"
                row["store_file_path"] = step2_store_path

        # 执行员工书名号引用 -> 目标门店sheet（第3步上传门店）
        emp_ref = _extract_book_title_ref(row.get("executor_employees", ""))
        if emp_ref:
            store_path = _save_sheet_blob_for_store(emp_ref, idx)
            row["upload_stores"] = "是"
            row["store_file_path"] = store_path
            # 会员通消息/朋友圈渠道下，执行员工仍为必填：用主消费营运区做兜底，避免仅书名号导致执行员工为空
            if any(p in {"会员通-发客户消息", "会员通-发客户朋友圈"} for p in parts):
                fallback_exec = str(row.get("main_operating_area", "") or "").strip()
                if fallback_exec:
                    row["executor_employees"] = fallback_exec

        # 客户消息/朋友圈：若未填“执行员工”，自动复用“主消费营运区”
        # 规则：
        # 1) xx大区/xx省区 -> 执行员工复用并默认同步加盟；
        # 2) 不含“大区/省区”视为营运区（如 大郑州营运区）-> 也默认同步加盟；
        # 3) 《xxx》在上方已转为上传门店逻辑。
        if (is_customer_msg or is_moments):
            exec_raw = str(row.get("executor_employees", "") or "").strip()
            main_area_raw = str(row.get("main_operating_area", "") or "").strip()
            if (not exec_raw) and main_area_raw:
                if _extract_book_title_ref(main_area_raw):
                    # 《xxx》场景：仅走上传门店，不强行写执行员工文本
                    pass
                else:
                    row["executor_employees"] = main_area_raw
                    row["executor_include_franchise"] = "是"

        # 购买目标商品编码书名号引用 -> 商品编码上传sheet
        product_ref = _extract_book_title_ref(row.get("purchase_target_product_code", ""))
        if product_ref:
            row["step2_product_file_path"] = _save_sheet_blob_for_step2_product(product_ref, idx)

        # 已领或已使用券规则ID书名号引用 -> 券规则ID文本（/拼接）
        coupon_ref = _extract_book_title_ref(row.get("coupon_ids_sheet_ref", ""))
        if coupon_ref:
            asset = sheet_assets.get(_norm_sheet_name(coupon_ref))
            if not asset:
                raise HTTPException(status_code=400, detail=f"第{idx}行：未找到sheet《{coupon_ref}》")
            _ensure_sheet_has_key_header(
                asset.get("rows", []),
                aliases=["券规则id", "券规则ID", "ruleid", "couponid"],
                row_no=idx,
                sheet_title=coupon_ref,
                field_desc="券规则ID",
            )
            values = _collect_sheet_first_col_values(asset.get("rows", []))
            if not values:
                raise HTTPException(status_code=400, detail=f"第{idx}行：sheet《{coupon_ref}》未找到有效券规则ID")
            row["coupon_ids"] = "/".join(values)

        # 小程序自动判定：有名称/标题/链接任一字段，自动开启添加小程序；
        # 若封面路径缺失，则自动复用最近一次已上传的封面图用于测试验证。
        mp_name = str(row.get("msg_mini_program_name", "") or "").strip()
        mp_title = str(row.get("msg_mini_program_title", "") or "").strip()
        mp_link = str(row.get("msg_mini_program_page_path", "") or "").strip()
        if mp_name or mp_title or mp_link:
            row["msg_add_mini_program"] = "是"
            if not mp_name:
                row["msg_mini_program_name"] = "大参林健康"
            if not str(row.get("msg_mini_program_cover_path", "") or "").strip():
                default_cover = _pick_default_mini_cover()
                if default_cover:
                    row["msg_mini_program_cover_path"] = default_cover

    with dst_csv.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        for row in normalized_rows:
            writer.writerow({k: row.get(k, "") for k in headers})


def convert_uploaded_xlsx_to_csv(upload: UploadFile, dst_csv: Path) -> None:
    if load_workbook is None:
        raise HTTPException(status_code=500, detail="Server missing openpyxl. Please install requirements-ui.txt")
    upload.file.seek(0)
    wb = load_workbook(upload.file, read_only=True, data_only=True)
    ws = wb.active
    with dst_csv.open("w", encoding="utf-8-sig", newline="") as out:
        writer = csv.writer(out)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(["" if v is None else str(v) for v in row])
    wb.close()


def _pick_sheet_name(sheet_names: List[str], candidates: List[str], fallback_keywords: List[str]) -> Optional[str]:
    """按候选名/关键字匹配sheet名称（忽略空格和大小写）。"""
    norm = {re.sub(r"\s+", "", n).lower(): n for n in sheet_names}
    for c in candidates:
        key = re.sub(r"\s+", "", c).lower()
        if key in norm:
            return norm[key]
    for name in sheet_names:
        n = re.sub(r"\s+", "", name).lower()
        if any(k in n for k in fallback_keywords):
            return name
    return None


def _write_sheet_to_csv(ws, dst_csv: Path) -> None:
    with dst_csv.open("w", encoding="utf-8-sig", newline="") as out:
        writer = csv.writer(out)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(["" if v is None else str(v) for v in row])


def _sheet_rows(ws) -> List[List[str]]:
    return [["" if v is None else str(v) for v in row] for row in ws.iter_rows(values_only=True)]


def _rows_to_dicts_with_norm_headers(rows: List[List[str]]) -> Tuple[List[str], List[Dict[str, str]]]:
    if not rows:
        return [], []
    raw_headers = [str(x or "").strip() for x in rows[0]]
    headers = [HEADER_CN_TO_EN.get(h, h) for h in raw_headers]
    out_rows: List[Dict[str, str]] = []
    for r in rows[1:]:
        d: Dict[str, str] = {}
        for i, h in enumerate(headers):
            d[h] = (str(r[i]).strip() if i < len(r) else "")
        if any(str(v or "").strip() for v in d.values()):
            out_rows.append(d)
    return headers, out_rows


def _sheet_to_xlsx_blob(ws_title: str, ws) -> Tuple[str, bytes]:
    if Workbook is None:
        raise HTTPException(status_code=500, detail="Server missing openpyxl. Please install requirements-ui.txt")
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = ws_title[:31] if ws_title else "Sheet1"
    for row in ws.iter_rows(values_only=True):
        out_ws.append(list(row))
    bio = io.BytesIO()
    out_wb.save(bio)
    out_wb.close()
    return f"{ws_title or 'sheet'}.xlsx", bio.getvalue()


def convert_uploaded_xlsx_multi_sheet_from_bytes(
    xlsx_bytes: bytes, dst_csv: Path
) -> Tuple[Optional[Tuple[str, bytes]], Optional[Tuple[str, bytes]], Dict[str, dict]]:
    """
    一个Excel多sheet模式：
    - 任务文件sheet -> 转CSV（非社群）
    - 社群任务sheet -> 追加转CSV（发送渠道=会员通-发送社群）
    - 目标门店sheet -> 返回xlsx blob
    - 目标商品sheet -> 返回xlsx blob
    """
    if load_workbook is None:
        raise HTTPException(status_code=500, detail="Server missing openpyxl. Please install requirements-ui.txt")
    wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    names = list(wb.sheetnames)

    task_sheet = _pick_sheet_name(
        names,
        candidates=["任务文件", "任务", "plans", "plan", "计划"],
        fallback_keywords=["任务", "plan", "plans", "计划"],
    ) or (names[0] if names else None)
    community_task_sheet = _pick_sheet_name(
        names,
        candidates=["社群任务", "社群任务文件", "community", "community_tasks"],
        fallback_keywords=["社群任务", "community"],
    )
    if community_task_sheet and task_sheet and (_norm_sheet_name(community_task_sheet) == _norm_sheet_name(task_sheet)):
        community_task_sheet = None
    if not task_sheet:
        wb.close()
        raise HTTPException(status_code=400, detail="Excel未找到可用sheet：任务文件")

    # 合并 任务文件 + 社群任务 到一个CSV
    task_headers, task_rows = _rows_to_dicts_with_norm_headers(_sheet_rows(wb[task_sheet]))
    community_rows: List[Dict[str, str]] = []
    community_headers: List[str] = []
    if community_task_sheet:
        community_headers, community_rows = _rows_to_dicts_with_norm_headers(_sheet_rows(wb[community_task_sheet]))
        for row in community_rows:
            if not str(row.get("channels", "") or "").strip():
                row["channels"] = "会员通-发送社群"

    final_headers = list(_default_headers())
    merged_rows = task_rows + community_rows
    for r in merged_rows:
        for k in r.keys():
            if k and (k not in final_headers):
                final_headers.append(k)
    with dst_csv.open("w", encoding="utf-8-sig", newline="") as out:
        writer = csv.DictWriter(out, fieldnames=final_headers)
        writer.writeheader()
        for r in merged_rows:
            writer.writerow({k: r.get(k, "") for k in final_headers})

    store_sheet = _pick_sheet_name(
        names,
        candidates=["目标门店", "门店", "主消费门店", "step2_main_store"],
        fallback_keywords=["目标门店", "门店"],
    )
    product_sheet = _pick_sheet_name(
        names,
        candidates=["目标商品", "商品编码", "商品", "step2_product"],
        fallback_keywords=["目标商品", "商品编码", "商品"],
    )

    store_blob: Optional[Tuple[str, bytes]] = None
    product_blob: Optional[Tuple[str, bytes]] = None
    sheet_assets: Dict[str, dict] = {}
    if store_sheet:
        store_blob = _sheet_to_xlsx_blob(store_sheet, wb[store_sheet])
    if product_sheet:
        product_blob = _sheet_to_xlsx_blob(product_sheet, wb[product_sheet])
    for n in names:
        ws = wb[n]
        blob_name, blob_bytes = _sheet_to_xlsx_blob(n, ws)
        rows = [["" if v is None else str(v) for v in row] for row in ws.iter_rows(values_only=True)]
        sheet_assets[_norm_sheet_name(n)] = {
            "title": n,
            "filename": blob_name,
            "bytes": blob_bytes,
            "rows": rows,
        }
    wb.close()
    return store_blob, product_blob, sheet_assets

def _is_valid_jpeg_png_bytes(data: bytes) -> bool:
    """校验图片真实格式（避免仅后缀正确但内容非法/加密）。"""
    if not data or len(data) < 16:
        return False
    # JPEG: starts with SOI(FFD8), and contains EOI(FFD9) later.
    # 不强制 EOI 在最后两个字节，避免被尾部扩展数据误伤。
    if data.startswith(b"\xff\xd8") and (b"\xff\xd9" in data[2:]):
        return True
    # PNG: signature + tail has IEND
    if data.startswith(b"\x89PNG\r\n\x1a\n") and (b"IEND" in data[-128:]):
        return True
    return False


def save_uploaded_moments_images(task_id: str, images: List[tuple[str, bytes]]) -> List[str]:
    """保存UI上传的朋友圈图片，返回本地绝对路径（按上传顺序）。"""
    if not images:
        return []
    out_dir = UPLOAD_DIR / f"{task_id}_images"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_paths: List[str] = []
    for idx, (name, data) in enumerate(images, 1):
        ext = Path(name).suffix.lower()
        if ext not in {".jpg", ".jpeg", ".png"}:
            raise HTTPException(status_code=400, detail=f"朋友圈图片格式仅支持 jpg/png: {name}")
        if len(data) >= 10 * 1024 * 1024:
            raise HTTPException(status_code=400, detail=f"朋友圈图片需小于10MB: {name}")
        safe = re.sub(r"[^0-9A-Za-z._-]+", "_", Path(name).name)
        dst = out_dir / f"{idx:02d}_{safe}"
        with dst.open("wb") as f:
            f.write(data)
        out_paths.append(str(dst.resolve()))
    if len(out_paths) > 9:
        raise HTTPException(status_code=400, detail=f"朋友圈图片最多9张，当前{len(out_paths)}张")
    return out_paths


def save_uploaded_mini_program_cover(task_id: str, image: tuple[str, bytes]) -> str:
    """保存UI上传的小程序封面，返回本地绝对路径。"""
    name, data = image
    out_dir = UPLOAD_DIR / f"{task_id}_mini_program"
    out_dir.mkdir(parents=True, exist_ok=True)
    ext = Path(name).suffix.lower()
    if ext not in {".jpg", ".jpeg", ".png"}:
        raise HTTPException(status_code=400, detail=f"小程序封面格式仅支持 jpg/png: {name}")
    if len(data) >= 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail=f"小程序封面需小于10MB: {name}")
    safe = re.sub(r"[^0-9A-Za-z._-]+", "_", Path(name).name)
    dst = out_dir / f"cover_{safe}"
    with dst.open("wb") as f:
        f.write(data)
    return str(dst.resolve())


def save_uploaded_store_file(task_id: str, store_file: tuple[str, bytes]) -> str:
    """保存UI上传的门店文件，返回本地绝对路径。"""
    name, data = store_file
    out_dir = UPLOAD_DIR / f"{task_id}_store_file"
    out_dir.mkdir(parents=True, exist_ok=True)
    ext = Path(name).suffix.lower()
    if ext not in {".xlsx", ".xls"}:
        raise HTTPException(status_code=400, detail=f"门店文件格式仅支持 xlsx/xls: {name}")
    safe = re.sub(r"[^0-9A-Za-z._-]+", "_", Path(name).name)
    dst = out_dir / f"store_{safe}"
    with dst.open("wb") as f:
        f.write(data)
    return str(dst.resolve())


def save_uploaded_main_store_file(task_id: str, store_file: tuple[str, bytes]) -> str:
    """保存第2步主消费门店上传文件，返回本地绝对路径。"""
    name, data = store_file
    out_dir = UPLOAD_DIR / f"{task_id}_step2_main_store"
    out_dir.mkdir(parents=True, exist_ok=True)
    ext = Path(name).suffix.lower()
    if ext not in {".xlsx", ".xls"}:
        raise HTTPException(status_code=400, detail=f"主消费门店文件格式仅支持 xlsx/xls: {name}")
    safe = re.sub(r"[^0-9A-Za-z._-]+", "_", Path(name).name)
    dst = out_dir / f"step2_main_store_{safe}"
    with dst.open("wb") as f:
        f.write(data)
    return str(dst.resolve())


def save_uploaded_step2_product_file(task_id: str, product_file: tuple[str, bytes]) -> str:
    """保存第2步商品编码上传文件，返回本地绝对路径（与门店文件分目录，避免覆盖）。"""
    name, data = product_file
    out_dir = UPLOAD_DIR / f"{task_id}_step2_product"
    out_dir.mkdir(parents=True, exist_ok=True)
    ext = Path(name).suffix.lower()
    if ext not in {".xlsx", ".xls"}:
        raise HTTPException(status_code=400, detail=f"第2步商品文件格式仅支持 xlsx/xls: {name}")
    safe = re.sub(r"[^0-9A-Za-z._-]+", "_", Path(name).name)
    dst = out_dir / f"step2_product_{safe}"
    with dst.open("wb") as f:
        f.write(data)
    return str(dst.resolve())


def inject_moments_images_to_csv(dst_csv: Path, image_paths: List[str], step3_channels: str) -> None:
    """将上传图片信息回写到任务CSV（覆盖朋友圈场景行）。"""
    if not image_paths:
        return
    with dst_csv.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        headers = list(reader.fieldnames or [])
    if not headers:
        return

    for col in ("moments_add_images", "moments_image_paths", "channels"):
        if col not in headers:
            headers.append(col)

    ui_channels = _normalize_channel_text(step3_channels or "")
    for row in rows:
        row_channels = _normalize_channel_text(str(row.get("channels", "") or "").strip())
        channel_scope = row_channels or ui_channels
        if ("会员通-发客户朋友圈" not in channel_scope) and ("会员通-发送社群" not in channel_scope):
            continue
        row["moments_add_images"] = "是"
        row["moments_image_paths"] = "|".join(image_paths)

    with dst_csv.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in headers})


def inject_message_mini_program_to_csv(
    dst_csv: Path,
    step3_channels: str,
    enabled: bool,
    program_name: str,
    title: str,
    cover_path: str,
    page_path: str,
) -> None:
    """将会员通消息类渠道（发客户消息/发送社群）的小程序配置回写到任务CSV。"""
    if not enabled:
        return
    with dst_csv.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        headers = list(reader.fieldnames or [])
    if not headers:
        return

    for col in (
        "msg_add_mini_program",
        "msg_mini_program_name",
        "msg_mini_program_title",
        "msg_mini_program_cover_path",
        "msg_mini_program_page_path",
        "channels",
    ):
        if col not in headers:
            headers.append(col)

    ui_channels = _normalize_channel_text(step3_channels or "")
    for row in rows:
        row_channels = _normalize_channel_text(str(row.get("channels", "") or "").strip())
        channel_scope = row_channels or ui_channels
        if ("会员通-发客户消息" not in channel_scope) and ("会员通-发送社群" not in channel_scope):
            continue
        row["msg_add_mini_program"] = "是"
        row["msg_mini_program_name"] = program_name or row.get("msg_mini_program_name", "") or "大参林健康"
        row["msg_mini_program_title"] = title or row.get("msg_mini_program_title", "") or ""
        row["msg_mini_program_cover_path"] = cover_path or ""
        row["msg_mini_program_page_path"] = page_path or row.get("msg_mini_program_page_path", "") or ""

    with dst_csv.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in headers})


def inject_store_file_to_csv(
    dst_csv: Path,
    step3_channels: str,
    enabled: bool,
    store_file_path: str,
) -> None:
    """将上传门店配置回写到任务CSV（会员通消息/朋友圈渠道）。"""
    if not enabled:
        return
    with dst_csv.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        headers = list(reader.fieldnames or [])
    if not headers:
        return

    for col in ("upload_stores", "store_file_path", "channels"):
        if col not in headers:
            headers.append(col)

    ui_channels = _normalize_channel_text(step3_channels or "")
    for row in rows:
        row_channels = _normalize_channel_text(str(row.get("channels", "") or "").strip())
        channel_scope = row_channels or ui_channels
        if ("会员通-发客户消息" not in channel_scope) and ("会员通-发送社群" not in channel_scope) and ("会员通-发客户朋友圈" not in channel_scope):
            continue
        row["upload_stores"] = "是"
        row["store_file_path"] = store_file_path or ""

    with dst_csv.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in headers})


def inject_step2_main_store_file_to_csv(
    dst_csv: Path,
    main_store_file_path: str,
) -> None:
    """将第2步主消费门店上传文件路径注入任务CSV。"""
    if not main_store_file_path:
        return
    with dst_csv.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        headers = list(reader.fieldnames or [])
    if not headers:
        return
    if "main_store_file_path" not in headers:
        headers.append("main_store_file_path")
    if "step2_store_file_path" not in headers:
        headers.append("step2_store_file_path")
    for row in rows:
        row["main_store_file_path"] = main_store_file_path
        row["step2_store_file_path"] = main_store_file_path
    with dst_csv.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in headers})


def inject_step2_product_file_to_csv(
    dst_csv: Path,
    product_file_path: str,
) -> None:
    """将第2步商品编码上传文件路径注入任务CSV。"""
    if not product_file_path:
        return
    with dst_csv.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        headers = list(reader.fieldnames or [])
    if not headers:
        return
    if "step2_product_file_path" not in headers:
        headers.append("step2_product_file_path")
    for row in rows:
        row["step2_product_file_path"] = product_file_path
    with dst_csv.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in headers})


def normalize_community_create_url_in_csv(dst_csv: Path, step3_channels: str) -> None:
    """
    社群任务创建链接标准化：
    - 若命中社群渠道且 create_url 为空 -> 填充 checkType=add 默认链接
    - 若命中社群渠道且 create_url 为旧 edit 链接 -> 转为 checkType=add
    """
    with dst_csv.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        headers = list(reader.fieldnames or [])
    if not headers:
        return
    if "create_url" not in headers:
        headers.append("create_url")
    if "channels" not in headers:
        headers.append("channels")

    default_add_url = "https://precision.dslyy.com/admin#/marketingPlan/addcommunityPlan?checkType=add"
    changed = False
    ui_channels = _normalize_channel_text(step3_channels or "")
    for row in rows:
        # 空白行不做任何默认链接注入，避免被误拆分为额外任务
        row_has_real_data = any(
            str(v or "").strip()
            for k, v in row.items()
            if k not in {"create_url"}
        )
        if not row_has_real_data:
            continue
        row_channels = _normalize_channel_text(str(row.get("channels", "") or "").strip())
        channel_scope = row_channels or ui_channels
        if "会员通-发送社群" not in channel_scope:
            continue
        old_url = str(row.get("create_url", "") or "").strip()
        new_url = old_url
        if not old_url:
            new_url = default_add_url
        elif "addcommunityPlan" in old_url and "checkType=edit" in old_url:
            new_url = re.sub(r"checkType=edit", "checkType=add", old_url)
        elif "addcommunityPlan" in old_url and "checkType=add" not in old_url:
            if "?" in old_url:
                new_url = f"{old_url}&checkType=add"
            else:
                new_url = f"{old_url}?checkType=add"
        if new_url != old_url:
            row["create_url"] = new_url
            changed = True

    if not changed:
        return
    with dst_csv.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in headers})


@dataclass
class TaskOptions:
    connect_cdp: bool = True
    cdp_endpoint: str = "http://127.0.0.1:18800"
    strict_step2: bool = True
    skip_step2: bool = False
    concurrent: int = 1
    start: Optional[int] = None
    end: Optional[int] = None
    hold_seconds: int = 2
    step3_channels: str = ""
    create_url: str = ""
    executor_include_franchise: bool = True
    notify_webhook: str = ""


@dataclass
class Task:
    id: str
    filename: str
    file_path: str
    options: TaskOptions
    operator: str = ""
    created_at: str = field(default_factory=now_iso)
    started_at: Optional[str] = None
    ended_at: Optional[str] = None
    status: str = "pending"
    total_plans: int = 0
    completed_plans: int = 0
    success_count: int = 0
    fail_count: int = 0
    eta: Optional[str] = None
    duration_sec: Optional[float] = None
    error: str = ""
    logs: List[str] = field(default_factory=list)
    generated_links: List[str] = field(default_factory=list)
    plan_name_display: str = ""
    channel_display: str = ""
    queued: bool = False
    paused: bool = False
    deleted: bool = False

    def _latest_link_for_ui(self) -> str:
        # 优先给业务展示真正可复核的 viewPlan 链接，其次 editPlan。
        for u in reversed(self.generated_links):
            if "#/marketingPlan/viewPlan?" in u:
                return u
        for u in reversed(self.generated_links):
            if "#/marketingPlan/editPlan?" in u:
                return u
        return self.generated_links[-1] if self.generated_links else ""

    def to_dict(self) -> dict:
        return {
            "id": self.id,
            "filename": self.filename,
            "file_path": self.file_path,
            "operator": self.operator,
            "created_at": self.created_at,
            "started_at": self.started_at,
            "ended_at": self.ended_at,
            "status": self.status,
            "total_plans": self.total_plans,
            "completed_plans": self.completed_plans,
            "success_count": self.success_count,
            "fail_count": self.fail_count,
            "eta": self.eta,
            "duration_sec": self.duration_sec,
            "error": self.error,
            "logs_count": len(self.logs),
            "generated_links": self.generated_links,
            "latest_link": self._latest_link_for_ui(),
            "plan_name": self.plan_name_display,
            "send_channels": self.channel_display,
            "queued": self.queued,
            "paused": self.paused,
            "options": {
                "connect_cdp": self.options.connect_cdp,
                "cdp_endpoint": self.options.cdp_endpoint,
                "strict_step2": self.options.strict_step2,
                "skip_step2": self.options.skip_step2,
                "concurrent": self.options.concurrent,
                "start": self.options.start,
                "end": self.options.end,
                "hold_seconds": self.options.hold_seconds,
                "step3_channels": self.options.step3_channels,
                "create_url": self.options.create_url,
                "executor_include_franchise": self.options.executor_include_franchise,
                "notify_webhook": self.options.notify_webhook,
            },
        }


def summarize_csv_meta(csv_path: Path) -> tuple[str, str]:
    names: List[str] = []
    channels: List[str] = []
    try:
        with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for i, row in enumerate(reader):
                n = (row.get("name", "") or "").strip()
                c = _normalize_channel_text((row.get("channels", "") or "").strip())
                if n and n not in names:
                    names.append(n)
                if c:
                    for p in re.split(r"[|,，、/]+", c):
                        p = (p or "").strip()
                        if p and p not in channels:
                            channels.append(p)
                if i >= 30:
                    break
    except Exception:
        return "", ""
    if not names:
        plan_name = ""
    elif len(names) == 1:
        plan_name = names[0]
    else:
        plan_name = f"{names[0]} +{len(names)-1}"
    return plan_name, "、".join(channels)


def parse_task_plans(csv_path: Path) -> List[dict]:
    plans: List[dict] = []
    try:
        with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for idx, row in enumerate(reader):
                name = str(row.get("name", "") or "").strip() or f"计划{idx+1}"
                channels = _normalize_channel_text(str(row.get("channels", "") or "").strip())
                plans.append(
                    {
                        "index": idx,
                        "name": name,
                        "channels": channels,
                        "plan_image_id": str(row.get("plan_image_id", "") or "").strip(),
                        "msg_add_mini_program": str(row.get("msg_add_mini_program", "") or "").strip() in {"是", "true", "True", "1"},
                        "moments_add_images": str(row.get("moments_add_images", "") or "").strip() in {"是", "true", "True", "1"},
                        "msg_mini_program_cover_path": str(row.get("msg_mini_program_cover_path", "") or "").strip(),
                        "moments_image_paths": str(row.get("moments_image_paths", "") or "").strip(),
                    }
                )
    except Exception:
        return plans
    return plans


def split_csv_to_single_plan_files(src_csv: Path, stem: str) -> List[Path]:
    """将一个任务CSV按计划行拆分为多个单计划CSV文件。"""
    with src_csv.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        headers = list(reader.fieldnames or [])
    if not headers:
        return [src_csv]

    def _row_has_plan_data(row: dict) -> bool:
        # 只把有业务内容的行视为计划，过滤 Excel 末尾空行/占位行
        keys = [
            "name",
            "channels",
            "region",
            "theme",
            "push_content",
            "send_time",
            "end_time",
        ]
        for k in keys:
            v = str(row.get(k, "") or "").strip()
            if v:
                return True
        # create_url / 文件路径等附属字段不单独作为“有效计划行”判定
        return False

    valid_rows = [r for r in rows if _row_has_plan_data(r)]
    if len(valid_rows) <= 1:
        return [src_csv]

    out_paths: List[Path] = []
    for i, row in enumerate(valid_rows, 1):
        tid = str(uuid.uuid4())
        out = UPLOAD_DIR / f"{tid}_{stem}_plan{i}.csv"
        with out.open("w", encoding="utf-8-sig", newline="") as fw:
            w = csv.DictWriter(fw, fieldnames=headers)
            w.writeheader()
            w.writerow({k: row.get(k, "") for k in headers})
        out_paths.append(out)
    return out_paths


def apply_task_materials_to_csv(csv_path: Path, specs: List[dict]) -> int:
    with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        headers = list(reader.fieldnames or [])
    if not headers:
        return 0

    required = [
        "msg_add_mini_program",
        "msg_mini_program_name",
        "msg_mini_program_title",
        "msg_mini_program_cover_path",
        "msg_mini_program_page_path",
        "moments_add_images",
        "moments_image_paths",
    ]
    for col in required:
        if col not in headers:
            headers.append(col)

    touched = 0
    spec_map = {}
    for s in specs:
        try:
            spec_map[int(s.get("index"))] = s
        except Exception:
            continue

    for idx, row in enumerate(rows):
        s = spec_map.get(idx)
        if not s:
            continue
        touched += 1
        if "msg_add_mini_program" in s:
            row["msg_add_mini_program"] = "是" if bool(s.get("msg_add_mini_program")) else "否"
        if s.get("msg_mini_program_name") is not None:
            row["msg_mini_program_name"] = str(s.get("msg_mini_program_name") or "")
        if s.get("msg_mini_program_title") is not None:
            row["msg_mini_program_title"] = str(s.get("msg_mini_program_title") or "")
        if s.get("msg_mini_program_page_path") is not None:
            row["msg_mini_program_page_path"] = str(s.get("msg_mini_program_page_path") or "")
        if s.get("msg_mini_program_cover_path") is not None:
            row["msg_mini_program_cover_path"] = str(s.get("msg_mini_program_cover_path") or "")

        if "moments_add_images" in s:
            row["moments_add_images"] = "是" if bool(s.get("moments_add_images")) else "否"
        if s.get("moments_image_paths") is not None:
            row["moments_image_paths"] = str(s.get("moments_image_paths") or "")

    with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in headers})
    return touched


def apply_plan_image_zip_to_csv(task_id: str, csv_path: Path, zip_bytes: bytes) -> Dict[str, int]:
    """
    批量图片包规则：
    - CSV 字段 plan_image_id（计划图片ID）为数字
    - zip 中子目录名需等于 plan_image_id，例如 1/2/3...
    - 子目录内文件名包含“小卡” => 小程序封面（社群/1对1）
    - 其余图片文件名按“文件名前缀数字”识别 1~9（如 1.jpg / 1-海报.jpg），按数字升序写入“添加图片”
    """
    with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        headers = list(reader.fieldnames or [])
    if not headers:
        raise HTTPException(status_code=400, detail="任务CSV无有效表头")

    for col in (
        "moments_add_images",
        "moments_image_paths",
        "plan_image_id",
        "msg_add_mini_program",
        "msg_mini_program_cover_path",
    ):
        if col not in headers:
            headers.append(col)

    group_moments: Dict[str, List[Tuple[int, str, bytes]]] = {}
    group_mini: Dict[str, Tuple[str, bytes]] = {}
    try:
        zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
    except Exception:
        raise HTTPException(status_code=400, detail="图片包格式错误，仅支持zip")

    allowed_ext = {".jpg", ".jpeg", ".png"}
    for info in zf.infolist():
        if info.is_dir():
            continue
        name = (info.filename or "").replace("\\", "/").strip("/")
        if not name:
            continue
        if name.startswith("__MACOSX/"):
            continue
        if "/._" in name or name.startswith("._"):
            continue
        parts = name.split("/")
        if len(parts) < 2:
            continue
        folder_raw = ""
        m_pid = None
        for seg in reversed(parts[:-1]):
            seg = (seg or "").strip()
            m = re.search(r"(\d+)", seg)
            if m:
                folder_raw = seg
                m_pid = m
                break
        if not m_pid:
            continue
        folder = str(int(m_pid.group(1)))
        base = Path(parts[-1]).stem.strip()
        ext = Path(parts[-1]).suffix.lower()
        if ext not in allowed_ext:
            continue
        blob = zf.read(info.filename)
        if not blob:
            continue
        m_num = re.match(r"^(\d+)(?:$|[^0-9].*)", base)

        # 小程序封面判定：
        # 1) 文件名包含“小卡”
        # 2) 或者文件名不以数字开头（兼容中文压缩包文件名编码异常导致“小卡”失真）
        if ("小卡" in base) or (m_num is None):
            # 同目录若有多个“小卡”，按文件名字典序取第一个，避免随机性
            old = group_mini.get(folder)
            if (old is None) or (parts[-1] < old[0]):
                group_mini[folder] = (parts[-1], blob)
            continue

        # 其余按“前缀数字”识别顺序，支持 1.jpg / 1-xxx.jpg
        n = int(m_num.group(1))
        if n < 1 or n > 9:
            continue
        group_moments.setdefault(folder, []).append((n, parts[-1], blob))

    matched = 0
    updated = 0
    updated_mini = 0
    updated_moments = 0
    for idx, row in enumerate(rows):
        pid = str(row.get("plan_image_id", "") or "").strip()
        if not pid:
            continue
        if not pid.isdigit():
            raise HTTPException(status_code=400, detail=f"第{idx+1}行：计划图片ID必须为数字，当前={pid}")
        imgs = group_moments.get(pid, [])
        mini = group_mini.get(pid)
        if (not imgs) and (not mini):
            continue
        matched += 1

        channels = [p.strip() for p in re.split(r"[|,，、/]+", str(row.get("channels", "") or "")) if p.strip()]
        channel_set = set(channels)
        support_mini = bool({"会员通-发客户消息", "会员通-发送社群"} & channel_set)
        support_moments = bool({"会员通-发客户消息", "会员通-发送社群", "会员通-发客户朋友圈"} & channel_set)

        if mini and support_mini:
            mini_name, mini_blob = mini
            mini_path = save_uploaded_mini_program_cover(
                f"{task_id}_pid{pid}_r{idx+1}",
                (mini_name, mini_blob),
            )
            if mini_path:
                row["msg_add_mini_program"] = "是"
                row["msg_mini_program_cover_path"] = mini_path
                updated = updated + 1
                updated_mini = updated_mini + 1

        if imgs and support_moments:
            imgs.sort(key=lambda x: x[0])
            nums = [n for n, _, _ in imgs]
            uniq = []
            for n in nums:
                if n not in uniq:
                    uniq.append(n)
            if len(uniq) != len(nums):
                raise HTTPException(status_code=400, detail=f"计划图片ID={pid} 子目录图片序号重复，请确保1~9唯一")
            if uniq != list(range(1, len(uniq) + 1)):
                raise HTTPException(status_code=400, detail=f"计划图片ID={pid} 子目录图片需按1开始连续命名（1,2,3...）")
            blobs = [(fname, b) for _, fname, b in imgs]
            out_paths = save_uploaded_moments_images(f"{task_id}_pid{pid}_r{idx+1}", blobs)
            if out_paths:
                row["moments_add_images"] = "是"
                row["moments_image_paths"] = "|".join(out_paths)
                updated = updated + 1
                updated_moments = updated_moments + 1

    with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in headers})
    return {
        "matched": matched,
        "updated": updated,
        "updated_mini": updated_mini,
        "updated_moments": updated_moments,
    }


def _parse_channel_list(raw: str) -> List[str]:
    return [p.strip() for p in re.split(r"[|,，、/]+", str(raw or "")) if p.strip()]


def _is_community_only_channels(raw: str) -> bool:
    ch = _parse_channel_list(raw)
    return bool(ch) and all(c == "会员通-发送社群" for c in ch)


class TaskRunner:
    def __init__(self, workers: int = 2):
        self.tasks: Dict[str, Task] = {}
        self.queue: asyncio.Queue[str] = asyncio.Queue()
        self.lock = asyncio.Lock()
        self.workers = max(1, workers)
        self.worker_tasks: List[asyncio.Task] = []

    async def start(self) -> None:
        if self.worker_tasks:
            return
        for i in range(self.workers):
            self.worker_tasks.append(asyncio.create_task(self._worker_loop(i + 1)))

    async def add_task(self, task: Task, auto_start: bool = True) -> None:
        async with self.lock:
            self.tasks[task.id] = task
        if auto_start:
            await self.enqueue_task(task.id)

    async def enqueue_task(self, task_id: str) -> bool:
        async with self.lock:
            task = self.tasks.get(task_id)
            if not task:
                return False
            if task.deleted or task.paused:
                return False
            if task.status != "pending":
                return False
            if task.queued:
                return False
            task.queued = True
        await self.queue.put(task_id)
        return True

    async def pause_task(self, task_id: str) -> Optional[Task]:
        async with self.lock:
            task = self.tasks.get(task_id)
            if not task or task.deleted:
                return None
            if task.status == "running":
                return task
            task.paused = True
            task.queued = False
            return task

    async def resume_task(self, task_id: str) -> Optional[Task]:
        async with self.lock:
            task = self.tasks.get(task_id)
            if not task or task.deleted:
                return None
            task.paused = False
            return task

    async def delete_task(self, task_id: str) -> bool:
        async with self.lock:
            task = self.tasks.get(task_id)
            if not task:
                return False
            if task.status == "running":
                return False
            task.deleted = True
            task.queued = False
            return True

    async def batch_pause_tasks(self, ids: List[str]) -> dict:
        changed = 0
        ignored = 0
        for tid in ids:
            t = await self.pause_task(tid)
            if t is None:
                ignored += 1
            elif t.status == "running":
                ignored += 1
            else:
                changed += 1
        return {"changed": changed, "ignored": ignored}

    async def batch_delete_tasks(self, ids: List[str]) -> dict:
        changed = 0
        ignored = 0
        for tid in ids:
            ok = await self.delete_task(tid)
            if ok:
                changed += 1
            else:
                ignored += 1
        return {"changed": changed, "ignored": ignored}

    async def retry_task(self, task_id: str) -> Task:
        async with self.lock:
            old = self.tasks.get(task_id)
            if not old or old.deleted:
                raise HTTPException(status_code=404, detail="Task not found")
            new_id = str(uuid.uuid4())
            new_task = Task(
                id=new_id,
                filename=old.filename,
                file_path=old.file_path,
                options=old.options,
                operator=old.operator,
            )
            self.tasks[new_id] = new_task
        await self.enqueue_task(new_id)
        return new_task

    async def start_pending(self) -> List[str]:
        async with self.lock:
            ids = [
                tid
                for tid, t in self.tasks.items()
                if t.status == "pending" and not t.queued and not t.paused and not t.deleted
            ]
        started: List[str] = []
        for tid in ids:
            ok = await self.enqueue_task(tid)
            if ok:
                started.append(tid)
        return started

    async def start_task(self, task_id: str) -> bool:
        return await self.enqueue_task(task_id)

    async def retry_failed(self) -> List[str]:
        async with self.lock:
            failed_ids = [tid for tid, t in self.tasks.items() if t.status == "failed" and not t.deleted]
        new_ids = []
        for tid in failed_ids:
            t = await self.retry_task(tid)
            new_ids.append(t.id)
        return new_ids

    async def list_tasks(self) -> List[dict]:
        async with self.lock:
            tasks = [t for t in self.tasks.values() if not t.deleted]
        tasks.sort(key=lambda x: x.created_at, reverse=True)
        return [t.to_dict() for t in tasks]

    async def get_task(self, task_id: str) -> Task:
        async with self.lock:
            t = self.tasks.get(task_id)
        if not t or t.deleted:
            raise HTTPException(status_code=404, detail="Task not found")
        return t

    async def append_log(self, task: Task, line: str) -> None:
        task.logs.append(line.rstrip("\n"))
        if len(task.logs) > 5000:
            task.logs = task.logs[-5000:]
        self._parse_progress(task, line)

    def _parse_progress(self, task: Task, line: str) -> None:
        m_total = re.search(r"计划数:\s*(\d+)", line)
        if m_total:
            task.total_plans = int(m_total.group(1))
        m_success = re.search(r"✅\s*成功:\s*(\d+)", line)
        if m_success:
            task.success_count = int(m_success.group(1))
        m_fail = re.search(r"❌\s*失败:\s*(\d+)", line)
        if m_fail:
            task.fail_count = int(m_fail.group(1))
        if "✅ 计划 " in line and " 完成！" in line:
            task.completed_plans += 1
        if "❌ 计划 " in line and " 失败 " in line:
            task.completed_plans += 1
        self._parse_generated_links(task, line)
        self._update_eta(task)

    def _parse_generated_links(self, task: Task, line: str) -> None:
        # Extract review links from runtime logs for business users.
        urls = re.findall(r"(https?://[^\s]+)", line)
        for u in urls:
            u = u.strip().rstrip(".,)")
            if "precision.dslyy.com" not in u:
                continue
            if (
                "#/marketingPlan/viewPlan?" in u
                or "#/marketingPlan/editPlan?" in u
                or "#/marketingTemplate/use?" in u
                or "useId=" in u
                or "#/marketingTemplate/" in u
            ):
                if u not in task.generated_links:
                    task.generated_links.append(u)
                    if len(task.generated_links) > 20:
                        task.generated_links = task.generated_links[-20:]

    def _update_eta(self, task: Task) -> None:
        if not task.started_at:
            return
        if task.total_plans <= 0 or task.completed_plans <= 0:
            task.eta = None
            return
        started = datetime.fromisoformat(task.started_at)
        elapsed = (datetime.now() - started).total_seconds()
        speed = task.completed_plans / max(elapsed, 1)
        remaining = max(task.total_plans - task.completed_plans, 0)
        eta = datetime.now() + timedelta(seconds=(remaining / max(speed, 1e-6)))
        task.eta = eta.isoformat(timespec="seconds")

    async def _notify_failure_if_needed(self, task: Task) -> None:
        webhook = (task.options.notify_webhook or "").strip()
        if not webhook:
            return
        summary = _extract_error_summary(task)
        tail = "\\n".join(_safe_tail(task.logs, 25))
        content = (
            f"【精准营销自动化】失败任务通知\\n"
            f"- 任务ID: {task.id}\\n"
            f"- 文件: {task.filename}\\n"
            f"- 计划: {task.plan_name_display or '-'}\\n"
            f"- 渠道: {task.channel_display or '-'}\\n"
            f"- 状态: {task.status}\\n"
            f"- 错误: {summary}\\n"
            f"- 开始: {task.started_at or '-'}\\n"
            f"- 结束: {task.ended_at or '-'}\\n"
            f"- 日志文件: {_task_history_log_path(task.id)}\\n"
            f"- 日志尾部:\\n{tail or '(空)'}"
        )
        data = json.dumps({"msg_type": "text", "content": {"text": content}}, ensure_ascii=False).encode("utf-8")
        req = urllib.request.Request(
            webhook,
            data=data,
            headers={"Content-Type": "application/json; charset=utf-8"},
            method="POST",
        )
        def _post() -> None:
            with urllib.request.urlopen(req, timeout=12) as resp:
                _ = resp.read()
        try:
            await asyncio.to_thread(_post)
            await self.append_log(task, "[notify] 飞书失败通知已发送")
        except Exception as e:
            await self.append_log(task, f"[notify] 飞书通知发送失败: {e}")

    async def _worker_loop(self, worker_id: int) -> None:
        while True:
            task_id = await self.queue.get()
            try:
                async with self.lock:
                    task = self.tasks.get(task_id)
                if not task or task.deleted:
                    continue
                if task.paused or task.status != "pending":
                    task.queued = False
                    continue
                # CDP 共享同一浏览器上下文时必须串行，避免多个任务互相抢页导致随机失败。
                if await self._has_other_running_cdp_task(task_id):
                    task.queued = True
                    await asyncio.sleep(1.0)
                    await self.queue.put(task_id)
                    continue
                await self._run_task(task, worker_id)
            finally:
                self.queue.task_done()

    async def _has_other_running_cdp_task(self, task_id: str) -> bool:
        async with self.lock:
            for tid, t in self.tasks.items():
                if tid == task_id or t.deleted:
                    continue
                if t.status == "running" and t.options.connect_cdp:
                    return True
        return False

    async def _run_task(self, task: Task, worker_id: int) -> None:
        task.status = "running"
        task.queued = False
        task.started_at = now_iso()
        task.error = ""
        task.logs = []
        await self.append_log(task, f"[worker-{worker_id}] task started: {task.filename}")

        # 双保险：worker侧再做一次社群自动策略兜底，避免上传侧策略未命中
        worker_community_only = _is_community_only_channels(task.options.step3_channels or task.channel_display)
        if worker_community_only:
            task.options.strict_step2 = False
            task.options.skip_step2 = True

        if task.options.skip_step2 and (not task.options.strict_step2):
            await self.append_log(task, "[worker-auto] 社群自动策略生效：已自动关闭严格第2步，并启用跳过第2步")

        cmd = [
            sys.executable,
            "-u",
            str(SCRIPT_PATH),
            "--csv",
            task.file_path,
            "--hold-seconds",
            str(task.options.hold_seconds),
        ]
        if task.options.connect_cdp:
            cmd.extend(["--connect-cdp", "--cdp-endpoint", task.options.cdp_endpoint])
        if task.options.strict_step2:
            cmd.append("--strict-step2")
        if task.options.skip_step2:
            cmd.append("--skip-step2")
        if task.options.concurrent:
            cmd.extend(["--concurrent", str(task.options.concurrent)])
        if task.options.start:
            cmd.extend(["--start", str(task.options.start)])
        if task.options.end:
            cmd.extend(["--end", str(task.options.end)])
        if task.options.step3_channels:
            cmd.extend(["--step3-channels", task.options.step3_channels])
        if task.options.create_url:
            cmd.extend(["--create-url", task.options.create_url])
        if task.options.executor_include_franchise:
            cmd.append("--executor-include-franchise")

        await self.append_log(task, f"$ {' '.join(cmd)}")
        started = datetime.now()
        child_env = os.environ.copy()
        # 避免在 nohup/后台启动场景中继承到无效标准流，导致 Python 子进程启动即崩溃
        child_env.setdefault("PYTHONUTF8", "1")
        child_env.setdefault("PYTHONIOENCODING", "utf-8")
        child_env.setdefault("PYTHONUNBUFFERED", "1")
        proc = await asyncio.create_subprocess_exec(
            *cmd,
            cwd=str(ROOT),
            env=child_env,
            stdin=asyncio.subprocess.DEVNULL,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.STDOUT,
        )
        assert proc.stdout is not None
        async for raw in proc.stdout:
            line = raw.decode("utf-8", errors="ignore")
            await self.append_log(task, line)
        rc = await proc.wait()

        task.ended_at = now_iso()
        task.duration_sec = (datetime.now() - started).total_seconds()
        if rc == 0 and task.fail_count == 0:
            task.status = "success"
        else:
            task.status = "failed"
            task.error = f"exit_code={rc}"
        await self.append_log(task, f"[worker-{worker_id}] task finished with status={task.status}")
        _persist_task_history(task)
        if task.status == "failed":
            _append_failure_index(task)
            await self._notify_failure_if_needed(task)
            _persist_task_history(task)


app = FastAPI(title="Precision Marketing Automation UI")
runner = TaskRunner(workers=2)


@app.on_event("startup")
async def startup_event() -> None:
    await runner.start()


@app.get("/", response_class=HTMLResponse)
async def index() -> str:
    return UI_HTML


@app.get("/api/tasks")
async def list_tasks() -> JSONResponse:
    return JSONResponse({"tasks": await runner.list_tasks()})


@app.get("/api/history/failed")
async def list_failed_history(limit: int = 100) -> JSONResponse:
    p = _task_failure_index_path()
    if not p.exists():
        return JSONResponse({"items": []})
    items: List[dict] = []
    try:
        lines = p.read_text(encoding="utf-8").splitlines()
        for ln in reversed(lines):
            if not ln.strip():
                continue
            try:
                items.append(json.loads(ln))
            except Exception:
                continue
            if len(items) >= max(1, min(limit, 500)):
                break
    except Exception:
        items = []
    return JSONResponse({"items": items})


@app.get("/api/history/task/{task_id}")
async def get_task_history(task_id: str) -> JSONResponse:
    p = _task_history_path(task_id)
    if not p.exists():
        raise HTTPException(status_code=404, detail="Task history not found")
    try:
        data = json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        raise HTTPException(status_code=500, detail="Task history parse failed")
    return JSONResponse(data)


@app.get("/api/tasks/{task_id}")
async def get_task(task_id: str) -> JSONResponse:
    task = await runner.get_task(task_id)
    return JSONResponse(task.to_dict())


@app.get("/api/tasks/{task_id}/file")
async def download_task_file(task_id: str):
    task = await runner.get_task(task_id)
    p = Path(task.file_path)
    if not p.exists():
        raise HTTPException(status_code=404, detail="Task file not found")
    return FileResponse(path=str(p), filename=p.name, media_type="text/csv")


@app.get("/api/tasks/{task_id}/plans")
async def get_task_plans(task_id: str) -> JSONResponse:
    task = await runner.get_task(task_id)
    p = Path(task.file_path)
    if not p.exists():
        raise HTTPException(status_code=404, detail="Task file not found")
    plans = parse_task_plans(p)
    return JSONResponse({"task_id": task_id, "plans": plans})


@app.get("/api/tasks/{task_id}/logs")
async def get_task_logs(task_id: str, offset: int = 0, limit: int = 300) -> JSONResponse:
    task = await runner.get_task(task_id)
    logs = task.logs[offset: offset + limit]
    return JSONResponse({
        "task_id": task_id,
        "offset": offset,
        "next_offset": offset + len(logs),
        "logs": logs,
        "status": task.status,
    })


@app.post("/api/tasks/{task_id}/materials")
async def save_task_materials(
    task_id: str,
    specs_json: str = Form(...),
    files: List[UploadFile] = File(default=[]),
) -> JSONResponse:
    task = await runner.get_task(task_id)
    csv_path = Path(task.file_path)
    if not csv_path.exists():
        raise HTTPException(status_code=404, detail="Task file not found")

    try:
        specs = json.loads(specs_json)
        if not isinstance(specs, list):
            raise ValueError("specs_json must be a list")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid specs_json: {e}")

    file_map: Dict[str, UploadFile] = {f.filename or "": f for f in files}
    out_dir = UPLOAD_DIR / f"{task_id}_plan_materials"
    out_dir.mkdir(parents=True, exist_ok=True)

    for spec in specs:
        idx = int(spec.get("index", -1))
        if idx < 0:
            continue
        plan_dir = out_dir / f"plan_{idx+1:03d}"
        plan_dir.mkdir(parents=True, exist_ok=True)

        cover_token = str(spec.get("cover_token", "") or "")
        if cover_token:
            uf = file_map.get(cover_token)
            if uf:
                b = await uf.read()
                if b:
                    cover_path = save_uploaded_mini_program_cover(f"{task_id}_p{idx+1}", (uf.filename or cover_token, b))
                    spec["msg_mini_program_cover_path"] = cover_path

        img_tokens = spec.get("moment_tokens", []) or []
        if img_tokens:
            img_blobs: List[Tuple[str, bytes]] = []
            for tok in img_tokens:
                tok = str(tok or "")
                uf = file_map.get(tok)
                if not uf:
                    continue
                b = await uf.read()
                if b:
                    img_blobs.append((uf.filename or tok, b))
            if img_blobs:
                img_paths = save_uploaded_moments_images(f"{task_id}_p{idx+1}", img_blobs)
                spec["moments_image_paths"] = "|".join(img_paths)

    touched = apply_task_materials_to_csv(csv_path, specs)
    return JSONResponse({"task_id": task_id, "updated_plans": touched})


@app.post("/api/tasks/{task_id}/materials/image-pack")
async def upload_task_image_pack(
    task_id: str,
    image_pack: UploadFile = File(...),
) -> JSONResponse:
    task = await runner.get_task(task_id)
    csv_path = Path(task.file_path)
    if not csv_path.exists():
        raise HTTPException(status_code=404, detail="Task file not found")
    if not image_pack.filename:
        raise HTTPException(status_code=400, detail="未选择图片包文件")
    if not image_pack.filename.lower().endswith(".zip"):
        raise HTTPException(status_code=400, detail="图片包仅支持 .zip")
    data = await image_pack.read()
    if not data:
        raise HTTPException(status_code=400, detail="图片包为空")

    stat = apply_plan_image_zip_to_csv(task_id, csv_path, data)
    return JSONResponse(
        {
            "task_id": task_id,
            "matched_plans": stat.get("matched", 0),
            "updated_plans": stat.get("updated", 0),
            "updated_mini": stat.get("updated_mini", 0),
            "updated_moments": stat.get("updated_moments", 0),
        }
    )


@app.get("/api/template/csv")
async def download_template_csv():
    p = UPLOAD_DIR / "精准营销任务模板_防乱码.csv"
    write_template_csv(p)
    return FileResponse(path=str(p), filename="精准营销任务模板（CSV防乱码）.csv", media_type="text/csv")


@app.get("/api/template/xlsx")
async def download_template_xlsx():
    if CUSTOM_EXPORT_TEMPLATE_XLSX.exists():
        return FileResponse(
            path=str(CUSTOM_EXPORT_TEMPLATE_XLSX),
            filename="精准营销任务模板（导入模板）.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    p = UPLOAD_DIR / "精准营销任务模板.xlsx"
    try:
        write_template_xlsx(p)
    except RuntimeError as e:
        raise HTTPException(status_code=500, detail=str(e))
    return FileResponse(
        path=str(p),
        filename="精准营销任务模板（导入模板）.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/api/template/community-xlsx")
async def download_community_template_xlsx():
    p = UPLOAD_DIR / "精准营销社群任务模板.xlsx"
    try:
        write_community_template_xlsx(p)
    except RuntimeError as e:
        raise HTTPException(status_code=500, detail=str(e))
    return FileResponse(
        path=str(p),
        filename="精准营销社群任务模板（含目标门店）.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.post("/api/tasks/upload")
async def upload_tasks(
    files: List[UploadFile] = File(...),
    moments_images: List[UploadFile] = File(default=[]),
    mini_program_cover: Optional[UploadFile] = File(default=None),
    store_file: Optional[UploadFile] = File(default=None),
    step2_main_store_file: Optional[UploadFile] = File(default=None),
    step2_product_file: Optional[UploadFile] = File(default=None),
    connect_cdp: bool = Form(True),
    cdp_endpoint: str = Form("http://127.0.0.1:18800"),
    strict_step2: bool = Form(True),
    skip_step2: bool = Form(False),
    concurrent: int = Form(1),
    start: str = Form(""),
    end: str = Form(""),
    hold_seconds: int = Form(2),
    notify_webhook: str = Form(""),
    step3_channels: str = Form(""),
    create_url: str = Form(""),
    executor_include_franchise: bool = Form(True),
    moments_add_images: bool = Form(False),
    upload_stores: bool = Form(False),
    msg_add_mini_program: bool = Form(False),
    msg_mini_program_name: str = Form("大参林健康"),
    msg_mini_program_title: str = Form(""),
    msg_mini_program_page_path: str = Form(""),
    operator: str = Form(""),
) -> JSONResponse:
    created = []
    created_task_ids: List[str] = []
    zip_files: List[UploadFile] = []
    task_files: List[UploadFile] = []
    options = TaskOptions(
        connect_cdp=connect_cdp,
        cdp_endpoint=cdp_endpoint.strip(),
        strict_step2=strict_step2,
        skip_step2=skip_step2,
        concurrent=max(1, concurrent),
        start=parse_int(start, 0) or None,
        end=parse_int(end, 0) or None,
        hold_seconds=max(0, hold_seconds),
        notify_webhook=notify_webhook.strip(),
        step3_channels=step3_channels.strip(),
        create_url=create_url.strip(),
        executor_include_franchise=executor_include_franchise,
    )

    # 支持一次上传多个文件：CSV/XLSX 任务文件 + ZIP 图片包
    for f in files:
        name = str(f.filename or "").lower()
        if name.endswith(".zip"):
            zip_files.append(f)
        elif name.endswith(".csv") or name.endswith(".xlsx"):
            task_files.append(f)
        else:
            raise HTTPException(status_code=400, detail=f"仅支持 CSV/XLSX/ZIP：{f.filename}")

    if not task_files:
        raise HTTPException(status_code=400, detail="请至少上传一个 CSV 或 XLSX 任务文件")

    # 朋友圈图片（由本UI上传），按上传顺序透传到任务CSV
    image_blobs: List[tuple[str, bytes]] = []
    if moments_add_images:
        for mf in moments_images:
            b = await mf.read()
            if b:
                image_blobs.append((mf.filename or f"image_{len(image_blobs)+1}.jpg", b))
        if not image_blobs:
            raise HTTPException(status_code=400, detail="已勾选朋友圈图片上传，但未选择图片文件")

    mini_cover_blob: Optional[tuple[str, bytes]] = None
    if msg_add_mini_program:
        if mini_program_cover is None:
            raise HTTPException(status_code=400, detail="已勾选添加小程序，但未上传小程序封面")
        b = await mini_program_cover.read()
        if not b:
            raise HTTPException(status_code=400, detail="小程序封面文件为空")
        mini_cover_blob = (mini_program_cover.filename or "mini_cover.jpg", b)

    store_file_blob: Optional[tuple[str, bytes]] = None
    if upload_stores:
        if store_file is None:
            # 允许走“单Excel多sheet”自动提取门店文件，无需前端单独上传
            store_file_blob = None
        else:
            b = await store_file.read()
            if not b:
                raise HTTPException(status_code=400, detail="门店文件为空")
            store_file_blob = (store_file.filename or "stores.xlsx", b)

    step2_main_store_blob: Optional[tuple[str, bytes]] = None
    if step2_main_store_file is not None:
        b = await step2_main_store_file.read()
        if b:
            step2_main_store_blob = (step2_main_store_file.filename or "step2_main_store.xlsx", b)
    step2_product_blob: Optional[tuple[str, bytes]] = None
    if step2_product_file is not None:
        b = await step2_product_file.read()
        if b:
            step2_product_blob = (step2_product_file.filename or "step2_product.xlsx", b)

    for f in task_files:
        lower = f.filename.lower()
        tid = str(uuid.uuid4())
        stem = Path(f.filename).stem
        dst = UPLOAD_DIR / f"{tid}_{stem}.csv"
        file_step2_store_blob: Optional[Tuple[str, bytes]] = None
        file_step2_product_blob: Optional[Tuple[str, bytes]] = None
        file_sheet_assets: Dict[str, dict] = {}
        if lower.endswith(".xlsx"):
            # 优先按“单Excel多sheet”读取；若无对应sheet则仅任务sheet生效
            raw_xlsx = await f.read()
            ms_store_blob, ms_product_blob, ms_sheet_assets = convert_uploaded_xlsx_multi_sheet_from_bytes(raw_xlsx, dst)
            file_step2_store_blob = ms_store_blob
            file_step2_product_blob = ms_product_blob
            file_sheet_assets = ms_sheet_assets or {}
        else:
            with dst.open("wb") as out:
                shutil.copyfileobj(f.file, out)
        normalize_uploaded_csv_headers(dst)
        normalize_channels_in_csv(dst)
        normalize_community_create_url_in_csv(dst, options.step3_channels)
        apply_unified_field_mapping_and_refs(dst, tid, options.step3_channels, file_sheet_assets)
        if moments_add_images and image_blobs:
            image_paths = save_uploaded_moments_images(tid, image_blobs)
            inject_moments_images_to_csv(dst, image_paths, options.step3_channels)
        if msg_add_mini_program and mini_cover_blob:
            mini_cover_path = save_uploaded_mini_program_cover(tid, mini_cover_blob)
            inject_message_mini_program_to_csv(
                dst,
                options.step3_channels,
                True,
                msg_mini_program_name.strip() or "大参林健康",
                msg_mini_program_title.strip(),
                mini_cover_path,
                msg_mini_program_page_path.strip(),
            )
        # 上传门店：优先显式上传文件；否则复用“第2步门店sheet/文件”
        resolved_store_blob = store_file_blob or step2_main_store_blob or file_step2_store_blob
        if upload_stores and resolved_store_blob:
            store_path = save_uploaded_store_file(tid, resolved_store_blob)
            inject_store_file_to_csv(
                dst,
                options.step3_channels,
                True,
                store_path,
            )
        resolved_step2_store_blob = step2_main_store_blob or file_step2_store_blob
        if resolved_step2_store_blob:
            step2_store_path = save_uploaded_main_store_file(tid, resolved_step2_store_blob)
            inject_step2_main_store_file_to_csv(dst, step2_store_path)
        resolved_step2_product_blob = step2_product_blob or file_step2_product_blob
        if resolved_step2_product_blob:
            step2_product_path = save_uploaded_step2_product_file(tid, resolved_step2_product_blob)
            inject_step2_product_file_to_csv(dst, step2_product_path)
        # 上传阶段前置时间校验：提前拦截“结束时间/发送时间小于当前时间”等错误
        prevalidate_csv_time_fields(dst)
        # 关键：一个上传文件内若有多条计划，拆成多条任务记录（每条计划一条任务）
        split_files = split_csv_to_single_plan_files(dst, stem)
        op = operator.strip() or os.getenv("USER") or getpass.getuser() or "unknown"
        for sf in split_files:
            plan_name_display, channel_display = summarize_csv_meta(sf)
            # 自动策略：仅社群渠道时，默认关闭严格第2步并启用跳过第2步（免人工配置）。
            # 优先使用任务文件中的渠道；若文件为空则回退到页面勾选渠道。
            community_only = _is_community_only_channels(channel_display or options.step3_channels)
            # 任务级优先使用文件内渠道；仅当文件没有渠道时才回退到页面全局渠道
            per_task_step3_channels = ("" if channel_display else options.step3_channels)
            file_options = TaskOptions(
                connect_cdp=options.connect_cdp,
                cdp_endpoint=options.cdp_endpoint,
                strict_step2=(False if community_only else options.strict_step2),
                skip_step2=(True if community_only else options.skip_step2),
                concurrent=options.concurrent,
                start=options.start,
                end=options.end,
                hold_seconds=options.hold_seconds,
                step3_channels=per_task_step3_channels,
                create_url=options.create_url,
                executor_include_franchise=options.executor_include_franchise,
            )
            task = Task(
                id=str(uuid.uuid4()),
                filename=f.filename if len(split_files) == 1 else f"{f.filename}#{sf.stem.split('_')[-1]}",
                file_path=str(sf),
                options=file_options,
                operator=op,
                plan_name_display=plan_name_display,
                channel_display=channel_display,
            )
            await runner.add_task(task, auto_start=False)
            created.append(task.to_dict())
            created_task_ids.append(task.id)

    # 若同批上传了图片ZIP包，则自动按“计划图片ID”写回每个新任务CSV
    zip_apply_stats: List[dict] = []
    if zip_files and created_task_ids:
        for zf in zip_files:
            zname = zf.filename or "images.zip"
            data = await zf.read()
            if (not data) or (not zname.lower().endswith(".zip")):
                continue
            for task_id in created_task_ids:
                task = await runner.get_task(task_id)
                csv_path = Path(task.file_path)
                if not csv_path.exists():
                    continue
                stat = apply_plan_image_zip_to_csv(task_id, csv_path, data)
                zip_apply_stats.append(
                    {
                        "task_id": task_id,
                        "zip": zname,
                        "matched_plans": stat.get("matched", 0),
                        "updated_plans": stat.get("updated", 0),
                        "updated_mini": stat.get("updated_mini", 0),
                        "updated_moments": stat.get("updated_moments", 0),
                    }
                )
    return JSONResponse({"created": created, "zip_applied": zip_apply_stats})


@app.post("/api/tasks/{task_id}/retry")
async def retry_task(task_id: str) -> JSONResponse:
    t = await runner.retry_task(task_id)
    return JSONResponse({"created": t.to_dict()})


@app.post("/api/tasks/{task_id}/pause")
async def pause_task(task_id: str) -> JSONResponse:
    t = await runner.pause_task(task_id)
    if not t:
        raise HTTPException(status_code=404, detail="Task not found")
    if t.status == "running":
        raise HTTPException(status_code=409, detail="运行中任务不支持暂停，请等待完成后再操作")
    return JSONResponse({"task": t.to_dict()})


@app.post("/api/tasks/{task_id}/resume")
async def resume_task(task_id: str) -> JSONResponse:
    t = await runner.resume_task(task_id)
    if not t:
        raise HTTPException(status_code=404, detail="Task not found")
    return JSONResponse({"task": t.to_dict()})


@app.post("/api/tasks/{task_id}/delete")
async def delete_task(task_id: str) -> JSONResponse:
    ok = await runner.delete_task(task_id)
    if not ok:
        raise HTTPException(status_code=409, detail="删除失败：任务不存在或正在运行")
    return JSONResponse({"task_id": task_id, "deleted": True})


@app.post("/api/tasks/pause-batch")
async def batch_pause_tasks(ids: List[str] = Body(...)) -> JSONResponse:
    if not isinstance(ids, list) or not ids:
        raise HTTPException(status_code=400, detail="ids 不能为空")
    result = await runner.batch_pause_tasks([str(i) for i in ids])
    return JSONResponse(result)


@app.post("/api/tasks/delete-batch")
async def batch_delete_tasks(ids: List[str] = Body(...)) -> JSONResponse:
    if not isinstance(ids, list) or not ids:
        raise HTTPException(status_code=400, detail="ids 不能为空")
    result = await runner.batch_delete_tasks([str(i) for i in ids])
    return JSONResponse(result)


@app.post("/api/tasks/start")
async def start_pending_tasks() -> JSONResponse:
    ids = await runner.start_pending()
    return JSONResponse({"started_ids": ids, "count": len(ids)})


@app.post("/api/tasks/{task_id}/start")
async def start_one_task(task_id: str) -> JSONResponse:
    ok = await runner.start_task(task_id)
    return JSONResponse({"task_id": task_id, "started": bool(ok)})


@app.post("/api/tasks/retry-failed")
async def retry_failed() -> JSONResponse:
    ids = await runner.retry_failed()
    return JSONResponse({"created_ids": ids})


UI_HTML = """
<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width,initial-scale=1" />
  <title>精准营销自动化任务中心</title>
  <style>
    :root{
      --bg:#f3f2fb;
      --card:#ffffff;
      --line:#ecebf5;
      --text:#1d1d1f;
      --sub:#515154;
      --hint:#6e6e73;
      --brand:#2f2a7e;
      --brand-dark:#221f5c;
      --brand-soft:#f1efff;
      --nav:#ffffff;
      --nav-active:#f4f2ff;
      --radius:16px;
      --control-h:40px;
      --font:14px;
      --space-1:8px;
      --space-2:12px;
      --space-3:16px;
      --space-4:20px;
    }
    *{box-sizing:border-box}
    body{font-family:"SF Pro Text","SF Pro Display","PingFang SC","Helvetica Neue",-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;margin:0;background:radial-gradient(1200px 700px at 20% -10%,#ede9ff 0%,transparent 60%),radial-gradient(1200px 700px at 100% 0%,#efe8ff 0%,transparent 58%),var(--bg);color:var(--text);font-size:var(--font);-webkit-font-smoothing:antialiased}
    .app-shell{max-width:1480px;margin:0 auto;min-height:100vh;padding:18px var(--space-3) 24px var(--space-3)}
    .main{padding:0}
    .page-title{margin:0 0 var(--space-2) 0;font-size:22px;line-height:1.2;font-weight:700;letter-spacing:-.01em;color:var(--text)}
    .card-title{margin:0 0 var(--space-2) 0;font-size:18px;line-height:1.25;font-weight:700;letter-spacing:-.01em;color:var(--text)}
    .card{background:rgba(255,255,255,.96);border:none;border-radius:18px;padding:var(--space-3);margin-bottom:var(--space-2);box-shadow:0 10px 30px rgba(42,34,94,.08)}
    .row{display:flex;gap:var(--space-2);flex-wrap:wrap;align-items:center}
    .section-title{font-size:15px;font-weight:700;color:#1d1d1f;margin:0 0 var(--space-2) 0;display:flex;align-items:center;gap:8px;line-height:1.3}
    .step-no{display:inline-flex;align-items:center;justify-content:center;width:22px;height:22px;border-radius:999px;background:var(--brand);color:#fff;font-size:12px;font-weight:700}
    .step-box{border:none;border-radius:14px;padding:var(--space-2);background:#fff;margin-bottom:10px;transition:all .2s ease;box-shadow:inset 0 0 0 1px rgba(120,110,190,.10)}
    .step-box:hover{box-shadow:inset 0 0 0 1px rgba(120,110,190,.16),0 8px 20px rgba(47,42,126,.06)}
    /* 按步骤区分渐变背景（参考卡片轻渐变风格） */
    .card .step-box:nth-of-type(1){
      background:linear-gradient(135deg,#f6f2ff 0%,#fff 55%,#f2f7ff 100%);
      border-color:transparent;
    }
    .card .step-box:nth-of-type(2){
      background:linear-gradient(135deg,#fff7f3 0%,#fff 52%,#f3f9ff 100%);
      border-color:transparent;
    }
    .card .step-box:nth-of-type(3){
      background:linear-gradient(135deg,#f2fff8 0%,#fff 55%,#f4f5ff 100%);
      border-color:transparent;
    }
    .step-box.compact{
      padding:10px;
      background:rgba(255,255,255,.52) !important;
      border:none !important;
      box-shadow:0 8px 22px rgba(62,54,120,.08) !important;
      backdrop-filter:blur(10px) saturate(118%);
      -webkit-backdrop-filter:blur(10px) saturate(118%);
    }
    .step-box.compact .form-grid{
      gap:8px 10px;
    }
    .step-box.compact .field{
      min-height:38px;
      gap:6px;
    }
    .step-box.compact .step-caption{
      margin-top:4px;
    }
    .step-caption{font-size:12px;color:var(--hint);margin-top:6px;line-height:1.6}
    .form-grid{display:grid;grid-template-columns:repeat(3,minmax(260px,1fr));gap:10px 12px}
    .form-grid .full{grid-column:1 / -1}
    .field{display:flex;align-items:center;gap:8px;min-height:44px}
    .field.between{justify-content:space-between}
    .field.vertical{flex-direction:column;align-items:flex-start}
    .label{min-width:112px;color:var(--sub);font-size:13px;font-weight:500}
    .inline-check{display:inline-flex;align-items:center;gap:6px;color:var(--sub);font-size:13px}
    .field input[type="text"], .field input[type="number"], .field input:not([type]), .field select{
      height:var(--control-h);box-sizing:border-box;
    }
    .field input[type="file"]{
      width:100%;
      max-width:560px;
      height:44px;
      padding:6px 10px;
      background:rgba(255,255,255,.78);
      border:1px solid rgba(120,110,190,.20);
      border-radius:12px;
      color:#66647a;
      line-height:30px;
    }
    .field input[type="file"]::file-selector-button{
      height:30px;
      padding:0 14px;
      margin-right:12px;
      border:1px solid #d0c9f2;
      border-radius:9px;
      background:linear-gradient(180deg,#f6f3ff,#ece7ff);
      color:#312b84;
      font-size:13px;
      font-weight:600;
      cursor:pointer;
    }
    .field input[type="file"]:hover{
      border-color:#c9c2ef;
      box-shadow:0 0 0 3px rgba(87,74,194,.08);
    }
    .field input[type="file"]::file-selector-button:hover{
      background:linear-gradient(180deg,#efeaff,#e4ddff);
    }
    /* 仅任务文件：去掉线框和背景，保持更简洁 */
    #files.file-uniform{
      background:transparent !important;
      border:none !important;
      box-shadow:none !important;
      padding:0 !important;
      height:auto !important;
      line-height:normal !important;
      color:var(--sub);
    }
    #files.file-uniform:hover{
      border:none !important;
      box-shadow:none !important;
    }
    #files.file-uniform::file-selector-button{
      margin-right:10px;
      border:none;
      border-radius:12px;
      background:linear-gradient(135deg,rgba(146,126,255,.26),rgba(134,189,255,.22));
      box-shadow:inset 0 0 0 1px rgba(114,126,196,.22);
    }
    .field.vertical .row{width:100%}
    .field.vertical .row label{display:flex;align-items:center;gap:6px;color:var(--sub);font-size:13px}
    .actions{display:flex;gap:10px;flex-wrap:wrap;align-items:center;padding-top:6px}
    .subcard{border:none;background:linear-gradient(180deg,#fefeff,#faf9ff);border-radius:12px;padding:12px;margin-top:8px;box-shadow:inset 0 0 0 1px rgba(120,110,190,.08)}
    .adv-toggle{display:inline-flex;align-items:center;gap:6px;height:34px;padding:0 14px;border-radius:999px;background:var(--brand-soft);color:#312b84;border:1px solid #dbd4ff;cursor:pointer;font-weight:600}
    .adv-toggle.text-link{
      height:auto;
      padding:0;
      border:none;
      background:transparent;
      color:#4b46a3;
      border-radius:0;
      box-shadow:none;
      font-weight:500;
      text-decoration:underline;
      text-underline-offset:3px;
    }
    .adv-toggle.text-link:hover{background:transparent;color:#312b84}
    .adv-panel{display:none;border:none;background:#faf8ff;border-radius:12px;padding:12px;margin-top:8px;box-shadow:inset 0 0 0 1px rgba(120,110,190,.14)}
    .adv-panel.open{display:block}
    .tiny{font-size:12px;color:var(--hint);line-height:1.55}
    .tiny.wrap{
      display:block;
      white-space:normal;
      word-break:normal;
      overflow-wrap:break-word;
      line-break:auto;
      max-width:100%;
    }
    .field.full .tiny{
      margin-left:8px;
      white-space:normal;
      overflow:visible;
      text-overflow:clip;
      max-width:none;
      line-height:1.5;
    }
    .upload-line{
      display:grid !important;
      grid-template-columns:112px minmax(560px,1fr) minmax(180px,auto);
      align-items:center;
      gap:12px;
      width:100%;
    }
    .upload-line .label{
      margin:0;
      min-width:112px;
      flex:none;
    }
    .upload-line .tiny{
      margin:0;
      max-width:none;
      white-space:normal;
      word-break:break-word;
    }
    .file-uniform{
      width:560px !important;
      max-width:560px !important;
      min-width:560px !important;
    }
    .channel-grid{display:grid;grid-template-columns:repeat(4,minmax(220px,1fr));gap:8px 14px}
    .compose-side .channel-grid{
      grid-template-columns:repeat(2,minmax(170px,1fr));
      gap:8px 10px;
    }
    .channel-block{border:none;border-radius:0;background:transparent;padding:0;box-shadow:none}
    .channel-item{
      display:flex;
      align-items:center;
      gap:10px;
      padding:6px 2px;
      border-radius:0;
      background:transparent;
      border:none;
      transition:all .2s ease;
    }
    .channel-item:hover{
      background:transparent;
      transform:none;
    }
    .channel-item input{margin-top:0;transform:scale(1.08)}
    .channel-icon{
      width:30px;
      height:30px;
      border-radius:10px;
      display:inline-flex;
      align-items:center;
      justify-content:center;
      font-size:12px;
      font-weight:800;
      letter-spacing:.2px;
      color:#fff;
      box-shadow:0 6px 12px rgba(28,35,52,.18);
      flex:0 0 30px;
    }
    .channel-icon.sms{background:linear-gradient(135deg,#8aa0c8 0%,#6f86b2 52%,#5d739f 100%)}
    .channel-icon.msg{background:linear-gradient(135deg,#334155 0%,#273449 48%,#1f2937 100%)}
    .channel-icon.group{background:linear-gradient(135deg,#1f9d8f 0%,#178f84 50%,#137a74 100%)}
    .channel-icon.moments{background:linear-gradient(135deg,#7b5cff 0%,#6b46f0 50%,#5b34dd 100%)}
    .channel-main{font-size:13px;color:#111827;font-weight:700;line-height:1.35}
    .channel-strong{font-weight:800;color:#312b84}
    .material-panel{border:1px solid #e8e6f3;background:#fcfcfe;border-radius:12px;padding:12px}
    .material-title{font-size:14px;font-weight:600;color:#0f172a;margin-bottom:8px}
    .channel-inline-config{margin-top:10px}
    .hidden{display:none !important}
    input,button,select{padding:8px 12px;border:1px solid #d7d4e8;border-radius:10px;font-size:13px;outline:none}
    input:focus,select:focus{border-color:#bdb5ed;box-shadow:0 0 0 3px rgba(87,74,194,.12)}
    button{background:linear-gradient(180deg,#3a328f,#2f2a7e);color:#fff;border:none;cursor:pointer;height:40px;padding:0 16px;font-weight:600;letter-spacing:.01em;box-shadow:0 6px 14px rgba(47,42,126,.24)}
    button:hover{background:linear-gradient(180deg,#302a82,#252066)}
    button.secondary{background:#fff;color:#2f2a7e;border:1px solid #cdc7ef;box-shadow:none}
    button.secondary:hover{background:#f7f5ff}
    .tip{font-size:12px;color:var(--hint);line-height:1.55}
    .hint{font-size:12px;color:var(--hint);display:block}
    table{width:100%;border-collapse:separate;border-spacing:0}
    th,td{border-bottom:1px solid rgba(120,110,190,.12);padding:10px 8px;text-align:left;font-size:13px;vertical-align:top}
    th{background:#f7f6fc;font-weight:600;color:#3e3a56;position:sticky;top:0;z-index:1}
    tbody tr:hover td{background:rgba(245,243,255,.45)}
    .status-pending{color:#6b7280}
    .status-running{color:#2563eb}
    .status-success{color:#059669}
    .status-failed{color:#dc2626}
    .status-paused{color:#b45309}
    #logs{background:#0b1020;color:#dbeafe;height:58vh;overflow:auto;padding:10px;border-radius:8px;white-space:pre-wrap;font-family:ui-monospace,Menlo,monospace;font-size:12px}
    .file-hero{
      background:transparent;
      border:none;
      border-radius:0;
      padding:0;
      display:flex;
      align-items:center;
      justify-content:space-between;
      gap:12px;
    }
    .file-hero .row{
      flex:1;
      min-width:0;
      align-items:center;
    }
    .file-hero .row .label{
      flex:0 0 112px;
    }
    .file-hero .row input[type="file"]{
      flex:1;
      max-width:none;
      min-width:260px;
    }
    .upload-actions{display:flex;gap:14px;align-items:center;margin-top:6px;padding-left:112px}
    .primary-actions button{height:44px;padding:0 20px;font-size:14px;font-weight:700}
    .log-modal{position:fixed;inset:0;background:rgba(15,23,42,.45);display:none;align-items:center;justify-content:center;z-index:9999}
    .log-modal.open{display:flex}
    .log-panel{width:min(1200px,92vw);max-height:88vh;background:#fff;border-radius:14px;border:1px solid #d1d5db;padding:14px}
    .log-head{display:flex;justify-content:space-between;align-items:center;margin-bottom:8px}
    .material-modal{position:fixed;inset:0;background:rgba(15,23,42,.45);display:none;align-items:center;justify-content:center;z-index:9998}
    .material-modal.open{display:flex}
    .material-panel{width:min(1260px,94vw);max-height:90vh;background:#fff;border-radius:14px;border:none;padding:14px;overflow:auto;box-shadow:0 18px 42px rgba(15,23,42,.16)}
    .material-row{border:none;border-radius:12px;padding:10px;margin-bottom:10px;box-shadow:inset 0 0 0 1px rgba(120,110,190,.10)}
    .material-row h4{margin:0 0 8px 0;font-size:14px}
    .material-grid{display:grid;grid-template-columns:1fr 1fr;gap:10px}
    .img-preview{display:flex;gap:8px;flex-wrap:wrap}
    .img-preview img{width:64px;height:64px;object-fit:cover;border-radius:8px;border:1px solid #ddd}
    .path-chip{display:inline-block;font-size:12px;padding:2px 8px;border:1px solid #ddd;border-radius:999px;background:#fafafa;margin:0 6px 6px 0}
    .link-pill{display:inline-block;padding:2px 8px;border-radius:999px;background:#f5f1ff;color:#312b84;border:1px solid #d9d0ff;font-size:12px;text-decoration:none}
    @media (max-width: 1100px){
      .form-grid{grid-template-columns:1fr}
      .label{min-width:90px}
      .channel-grid{grid-template-columns:1fr}
      .upload-line{grid-template-columns:1fr}
      .file-uniform{width:100% !important;min-width:0 !important;max-width:none !important}
    }
    /* === Editorial high-end visual refresh (DESIGN.md aligned, no behavior changes) === */
    @import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@500;600;700;800&family=Work+Sans:wght@400;500;600;700&display=swap');
    :root{
      --im-bg:#f9f9f9;
      --im-bg-soft:#f4f3f3;
      --im-bg-deep:#eceaea;
      --im-ink:#1c2334;
      --im-sub:#454b5a;
      --im-hint:#6a6f7d;
      --im-brand:#1c2334;
      --im-brand-2:#2b3348;
      --im-accent:#e4c02f;
      --im-accent-deep:#715d00;
      --im-card:#ffffff;
      --im-ghost:rgba(28,35,52,.13);
      --im-shadow:0 28px 56px rgba(28,35,52,.07);
    }
    body{
      background:
        radial-gradient(1200px 640px at 4% -8%, rgba(113,61,255,.18), transparent 58%),
        radial-gradient(1200px 640px at 96% -12%, rgba(60,96,255,.14), transparent 56%),
        linear-gradient(180deg,#f7f8fe 0%,#f1f3fb 100%);
      color:var(--im-ink);
      font-family:"Plus Jakarta Sans","Work Sans","PingFang SC","Helvetica Neue",-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;
    }
    .app-shell{
      max-width:1440px;
      padding:22px 18px 28px;
    }
    .card{
      background:linear-gradient(180deg,rgba(255,255,255,.72),rgba(246,248,255,.68));
      backdrop-filter:blur(16px) saturate(118%);
      -webkit-backdrop-filter:blur(16px) saturate(118%);
      border:1px solid rgba(113,61,255,.12);
      border-radius:22px;
      box-shadow:0 18px 42px rgba(48,56,92,.12);
      padding:18px;
      margin-bottom:14px;
    }
    .card-title{
      font-family:"Plus Jakarta Sans","PingFang SC","Helvetica Neue",-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;
      font-size:1.78rem;
      font-weight:700;
      color:var(--im-ink);
      margin-bottom:14px;
      letter-spacing:.2px;
    }
    .section-title{
      font-family:"Plus Jakarta Sans","PingFang SC","Helvetica Neue",-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;
      font-size:1.04rem;
      font-weight:700;
      color:var(--im-ink);
      margin-bottom:10px;
    }
    .step-no{
      width:24px;
      height:24px;
      background:linear-gradient(135deg,var(--im-brand),var(--im-brand-2));
      box-shadow:0 6px 14px rgba(28,35,52,.25);
    }
    .step-box{
      border:1px solid rgba(113,61,255,.10);
      box-shadow:0 8px 20px rgba(54,58,98,.08);
      border-radius:16px;
      padding:12px;
      margin-bottom:12px;
      backdrop-filter:blur(8px) saturate(114%);
      -webkit-backdrop-filter:blur(8px) saturate(114%);
    }
    .card .step-box:nth-of-type(1){
      background:linear-gradient(135deg,#ffffff 0%,#f4f3f3 100%);
    }
    .card .step-box:nth-of-type(2){
      background:linear-gradient(135deg,#fbfbfb 0%,#f2f2f2 100%);
    }
    .card .step-box:nth-of-type(3){
      background:linear-gradient(135deg,#fefefe 0%,#efefef 100%);
    }
    .step-caption,.tip,.tiny,.hint{
      color:var(--im-hint);
      line-height:1.55;
    }
    .label{
      color:var(--im-sub);
      font-weight:600;
    }
    .upload-line{
      grid-template-columns:104px minmax(580px,1fr) minmax(180px,auto);
      gap:10px;
    }
    input,button,select{
      border-radius:10px;
      border:1px solid rgba(28,35,52,.10);
      font-size:13px;
    }
    input[type="text"],input[type="number"],select{
      background:transparent;
      border:none;
      border-bottom:1px solid rgba(28,35,52,.24);
      border-radius:0;
      padding-left:0;
      padding-right:0;
    }
    input:focus,select:focus{
      border-color:var(--im-brand);
      box-shadow:none;
    }
    .field input[type="file"]{
      height:42px;
      padding:5px 10px;
      border:1px solid rgba(28,35,52,.16);
      border-radius:12px;
      background:linear-gradient(180deg,#ffffff,#f4f3f3);
      color:#3f4657;
    }
    .field input[type="file"]::file-selector-button{
      border-radius:10px;
      border:1px solid rgba(28,35,52,.18);
      background:linear-gradient(180deg,#ffffff,#eeeeef);
      color:#222938;
      font-weight:600;
    }
    .field input[type="file"]::file-selector-button:hover{
      background:linear-gradient(180deg,#f8f8f8,#e9e9ea);
    }
    .adv-panel{
      background:linear-gradient(180deg,#fbfbfb,#f3f3f4);
      border:none;
      box-shadow:none;
    }
    .adv-toggle.text-link{
      color:#2f3650;
      text-decoration:none;
      border-bottom:1px dashed rgba(28,35,52,.38);
      font-weight:600;
    }
    .adv-toggle.text-link:hover{
      color:#111827;
      border-bottom-color:rgba(28,35,52,.62);
    }
    button{
      background:linear-gradient(135deg,#7c89d9 0%,#9aa5ea 52%,#b7c0f2 100%);
      border:1px solid rgba(92,108,185,.38);
      box-shadow:0 10px 22px rgba(124,137,217,.24);
      font-weight:700;
      letter-spacing:.15px;
      color:#ffffff;
    }
    button:hover{
      background:linear-gradient(135deg,#8896e1 0%,#a8b2ee 54%,#c3cbf6 100%);
      border-color:rgba(92,108,185,.44);
    }
    button.secondary{
      background:linear-gradient(135deg,#f8faff 0%,#eef2ff 50%,#f5f7ff 100%);
      color:#4a5685;
      border:1px solid rgba(123,137,217,.30);
      box-shadow:none;
      font-weight:600;
    }
    button.secondary:hover{
      background:linear-gradient(135deg,#f3f6ff 0%,#e9eeff 52%,#f1f4ff 100%);
      border-color:rgba(123,137,217,.42);
      color:#3f4b79;
    }
    .primary-actions button{
      min-width:140px;
      height:46px;
    }
    .link-pill{
      border:1px solid rgba(28,35,52,.18);
      background:linear-gradient(180deg,#fbfbfb,#f2f2f2);
      color:#2b3348;
      font-weight:600;
      border-radius:999px;
      padding:4px 10px;
    }
    table{
      border:none;
      border-radius:12px;
      overflow:hidden;
      background:linear-gradient(180deg,#ffffff,#f8f8f8);
    }
    th{
      background:linear-gradient(180deg,#f2f2f3,#ececec);
      color:#2d3342;
      font-weight:700;
      border-bottom:1px solid rgba(28,35,52,.08);
    }
    td{
      border-bottom:1px solid rgba(28,35,52,.06);
      color:#1f2937;
    }
    tbody tr:hover td{
      background:rgba(28,35,52,.035);
    }
    .material-panel,.log-panel{
      border:none;
      border-radius:16px;
      box-shadow:0 22px 52px rgba(28,35,52,.22);
    }
    .material-row{
      border:none;
      box-shadow:none;
      background:linear-gradient(180deg,#ffffff,#f3f3f3);
    }
    .hero-bar{
      display:flex;
      align-items:flex-end;
      justify-content:space-between;
      gap:12px;
      margin-bottom:10px;
    }
    .hero-title{
      font-family:"Plus Jakarta Sans","PingFang SC","Helvetica Neue",-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;
      font-size:1.42rem;
      font-weight:700;
      color:var(--im-ink);
      letter-spacing:.1px;
    }
    .hero-sub{
      font-size:12px;
      color:var(--im-hint);
      margin-top:2px;
    }
    .theme-wrap{display:flex;align-items:center;gap:8px}
    .theme-switch{
      display:inline-flex;
      align-items:center;
      gap:4px;
      padding:4px;
      border:1px solid rgba(148,163,184,.28);
      border-radius:14px;
      background:linear-gradient(180deg,rgba(255,255,255,.96),rgba(244,247,252,.92));
      box-shadow:inset 0 1px 0 rgba(255,255,255,.86), 0 2px 8px rgba(15,23,42,.06);
    }
    .theme-btn{
      width:32px;height:32px;border-radius:10px;border:none;cursor:pointer;
      display:inline-flex;align-items:center;justify-content:center;
      color:#64748b;background:transparent;box-shadow:none;padding:0;
      transition:all .18s ease;
    }
    .theme-btn svg{width:16px;height:16px;stroke:currentColor;fill:none;stroke-width:1.8;stroke-linecap:round;stroke-linejoin:round}
    .theme-btn:hover{color:#111827;background:rgba(148,163,184,.16)}
    .theme-btn.active{
      background:linear-gradient(180deg,#ffffff,#f8fafc);
      color:#334155;
      box-shadow:inset 0 1px 0 rgba(255,255,255,.9), 0 1px 3px rgba(15,23,42,.12);
    }
    .compose-layout{
      display:block;
      margin-top:8px;
    }
    .upload-zone{
      border:2px dashed rgba(124,108,225,.32);
      background:rgba(255,255,255,.34);
      border-radius:24px;
      min-height:150px;
      display:flex;
      align-items:center;
      justify-content:center;
      text-align:center;
      cursor:pointer;
      transition:all .2s ease;
      margin-bottom:14px;
      padding:14px 16px;
    }
    .upload-zone:hover{
      border-color:rgba(113,61,255,.54);
      background:rgba(255,255,255,.45);
      box-shadow:inset 0 0 0 1px rgba(113,61,255,.1);
    }
    .upload-zone.drag-over{
      border-color:rgba(113,61,255,.72);
      background:rgba(255,255,255,.54);
      box-shadow:0 0 0 4px rgba(113,61,255,.12);
    }
    .upload-zone .up-icon{
      width:44px;height:44px;border-radius:14px;
      margin:0 auto 8px auto;
      display:flex;align-items:center;justify-content:center;
      color:#2f3c56;
      background:linear-gradient(135deg,rgba(113,61,255,.1),rgba(59,130,246,.1));
      box-shadow:inset 0 0 0 1px rgba(113,61,255,.16);
      font-size:20px;font-weight:800;
    }
    .upload-zone .up-title{
      font-size:20px;
      line-height:1.15;
      letter-spacing:-.02em;
      color:#111827;
      font-weight:800;
    }
    .upload-zone .up-sub{
      margin-top:6px;
      font-size:12px;
      color:#64748b;
      font-weight:500;
    }
    .config-foot{
      display:flex;
      align-items:center;
      justify-content:space-between;
      gap:10px;
      flex-wrap:wrap;
      margin:6px 0 8px 0;
    }
    .config-foot .left,
    .config-foot .right{display:flex;align-items:center;gap:10px;flex-wrap:wrap}
    .action-strip{
      display:flex;
      align-items:center;
      justify-content:space-between;
      gap:12px;
      margin:2px 0 10px 0;
      padding:2px 0;
      flex-wrap:wrap;
    }
    .action-left{display:flex;align-items:center;gap:12px;flex-wrap:wrap}
    .action-right{display:flex;align-items:center}
    .action-strip button{
      min-width:180px;
      height:48px;
      font-size:16px;
      border-radius:14px;
    }
    .action-strip button.secondary{
      min-width:200px;
    }
    .batch-op{
      display:flex;
      align-items:center;
      gap:10px;
      margin-left:0;
      padding:8px 10px;
      border:1px solid rgba(28,35,52,.12);
      border-radius:999px;
      background:rgba(255,255,255,.56);
    }
    .batch-op .tiny{
      margin:0;
      color:#6b7280;
      font-size:12px;
      font-weight:700;
      letter-spacing:.2px;
      padding-left:12px;
      margin-left:2px;
      border-left:1px solid rgba(28,35,52,.16);
      line-height:1;
      white-space:nowrap;
    }
    .batch-op button.secondary{
      min-width:110px;
      height:40px;
      padding:0 16px;
      border-radius:999px;
      font-size:14px;
      font-weight:700;
      background:rgba(255,255,255,.9);
      border:1px solid rgba(107,114,128,.22);
      box-shadow:inset 0 1px 0 rgba(255,255,255,.9), 0 1px 2px rgba(15,23,42,.04);
    }
    .batch-op button.secondary:hover{
      background:#fff;
      border-color:rgba(71,85,105,.32);
    }
    .task-table-wrap{
      border-radius:14px;
      overflow:hidden;
      background:#fff;
    }
    .check-col{width:38px;text-align:center}
    .check-col input{transform:scale(1.05)}
    .task-card-head{
      display:flex;
      align-items:center;
      justify-content:space-between;
      margin-bottom:10px;
    }
    .task-meta{
      font-size:12px;
      color:var(--im-hint);
    }
    .task-section-spacer{
      margin-top:2px;
      border-top:1px solid rgba(28,35,52,.08);
      padding-top:10px;
    }
    @media (max-width: 1180px){
      .compose-layout{grid-template-columns:1fr}
      .hero-bar{flex-direction:column;align-items:flex-start}
      .theme-wrap{width:100%}
      .theme-switch{width:100%;justify-content:flex-start}
    }
    html[data-theme="dark"]{color-scheme:dark}
    html[data-theme="dark"] body{
      background:
        radial-gradient(1200px 640px at 6% -12%, rgba(113,61,255,.22), transparent 60%),
        radial-gradient(1400px 760px at 100% -18%, rgba(58,92,201,.20), transparent 62%),
        linear-gradient(180deg,#21283a 0%,#1b2232 54%,#161d2b 100%);
      color:#ecf1fa;
    }
    html[data-theme="dark"] .card{
      background:linear-gradient(180deg,rgba(44,54,76,.68),rgba(35,44,63,.70));
      border:1px solid rgba(169,190,255,.16);
      backdrop-filter:blur(16px) saturate(122%);
      -webkit-backdrop-filter:blur(16px) saturate(122%);
      box-shadow:0 20px 48px rgba(12,18,34,.26), inset 0 1px 0 rgba(255,255,255,.07);
    }
    html[data-theme="dark"] .hero-title,
    html[data-theme="dark"] .section-title,
    html[data-theme="dark"] .card-title{color:#f5f8ff}
    html[data-theme="dark"] .label,
    html[data-theme="dark"] .channel-main{color:#dbe4f2}
    html[data-theme="dark"] .tiny,
    html[data-theme="dark"] .hint,
    html[data-theme="dark"] .step-caption,
    html[data-theme="dark"] .hero-sub,
    html[data-theme="dark"] .task-meta{color:#aab6c9}
    html[data-theme="dark"] .step-box{
      background:linear-gradient(180deg,#313b52,#283247) !important;
      box-shadow:inset 0 0 0 1px rgba(183,196,228,.14), 0 8px 20px rgba(15,21,35,.16);
    }
    html[data-theme="dark"] .step-box.compact{
      background:rgba(57,69,95,.52) !important;
      border:none !important;
      box-shadow:0 10px 24px rgba(8,13,24,.22) !important;
      backdrop-filter:blur(10px) saturate(120%);
      -webkit-backdrop-filter:blur(10px) saturate(120%);
    }
    html[data-theme="dark"] .field input[type="file"]{
      background:linear-gradient(180deg,#334059,#2a3448);
      border-color:rgba(203,216,245,.22);
      color:#e8eefc;
    }
    html[data-theme="dark"] .field input[type="file"]::file-selector-button{
      background:linear-gradient(180deg,#4a5a79,#394a67);
      border-color:rgba(203,216,245,.28);
      color:#f6f9ff;
    }
    html[data-theme="dark"] #files.file-uniform{
      background:transparent !important;
      border:none !important;
      box-shadow:none !important;
      color:#dbe4f3;
    }
    html[data-theme="dark"] #files.file-uniform:hover{
      border:none !important;
      box-shadow:none !important;
    }
    html[data-theme="dark"] #files.file-uniform::file-selector-button{
      background:linear-gradient(135deg,rgba(161,178,255,.28),rgba(124,206,255,.22));
      border:none;
      box-shadow:inset 0 0 0 1px rgba(214,226,246,.24);
      color:#f5f8ff;
    }
    html[data-theme="dark"] .adv-panel{background:linear-gradient(180deg,#303b54,#273349)}
    html[data-theme="dark"] .link-pill{
      background:linear-gradient(180deg,#394865,#2f3d58);
      border-color:rgba(203,216,245,.25);
      color:#e0e8ff;
    }
    html[data-theme="dark"] .task-table-wrap{background:#2a3449}
    html[data-theme="dark"] table{background:linear-gradient(180deg,#34425d,#2a3449)}
    html[data-theme="dark"] th{
      background:linear-gradient(180deg,#455677,#374762);
      color:#edf2fb;
      border-bottom-color:rgba(214,226,246,.24);
    }
    html[data-theme="dark"] td{
      color:#f1f5ff;
      border-bottom-color:rgba(214,226,246,.16);
    }
    html[data-theme="dark"] tbody tr:hover td{background:rgba(189,206,242,.12)}
    html[data-theme="dark"] .theme-switch{
      background:linear-gradient(180deg,#3a4967,#2f3d58);
      border-color:rgba(214,226,246,.26);
    }
    html[data-theme="dark"] .theme-btn{color:#b7c3d8}
    html[data-theme="dark"] .theme-btn:hover{color:#f4f7ff;background:rgba(214,226,246,.14)}
    html[data-theme="dark"] .theme-btn.active{
      background:linear-gradient(135deg,#8ea2ff 0%,#9f89ff 52%,#8de1ff 100%);
      color:#ffffff;
    }
    #logs{
      background:#132138;
      color:#e3ecff;
      border-radius:10px;
      border:1px solid rgba(180,202,243,.30);
    }

    /* ---- Design Overlay (non-breaking visual override) ---- */
    :root{
      --ds-accent:#713DFF;
      --ds-accent-glow:rgba(113,61,255,.18);
      --ds-glass-bg:rgba(255,255,255,.62);
      --ds-glass-border:rgba(255,255,255,.84);
      --ds-text:#1e293b;
      --ds-sub:#64748b;
    }
    body{
      font-family:"Plus Jakarta Sans","PingFang SC","Helvetica Neue",-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;
      background:
        radial-gradient(at 0% 0%, rgba(113,61,255,.08) 0px, transparent 50%),
        radial-gradient(at 100% 0%, rgba(59,130,246,.06) 0px, transparent 52%),
        radial-gradient(at 50% 100%, #f8fafc 0px, transparent 55%),
        #f4f6fb;
      color:var(--ds-text);
    }
    .app-shell{max-width:1440px;padding:28px 22px 24px}
    .card{
      background:var(--ds-glass-bg);
      border:1px solid var(--ds-glass-border);
      border-radius:24px;
      backdrop-filter:blur(20px) saturate(118%);
      -webkit-backdrop-filter:blur(20px) saturate(118%);
      box-shadow:0 10px 32px rgba(31,38,135,.08);
    }
    .hero-title{
      font-size:2rem;
      font-weight:800;
      letter-spacing:-.02em;
      background:linear-gradient(135deg,#1e293b 0%,#64748b 100%);
      -webkit-background-clip:text;
      -webkit-text-fill-color:transparent;
    }
    .hero-sub{color:var(--ds-sub);font-weight:500}
    .section-title,.card-title,.label,.channel-main{color:#1f2937}
    .tiny,.hint,.step-caption,.task-meta{color:#6b7280}
    .step-box{
      background:rgba(255,255,255,.72) !important;
      border:1px solid rgba(255,255,255,.88) !important;
      border-radius:18px;
      box-shadow:0 8px 26px rgba(31,38,135,.07);
    }
    .step-box.compact{
      background:rgba(255,255,255,.56) !important;
      border:1px solid rgba(255,255,255,.82) !important;
      box-shadow:0 8px 20px rgba(31,38,135,.06) !important;
    }
    input,select,button{
      border-radius:14px;
      transition:all .2s cubic-bezier(.4,0,.2,1);
    }
    input[type="text"], input[type="number"], input:not([type]), select{
      background:rgba(255,255,255,.72);
      border:1px solid rgba(0,0,0,.06);
      color:#334155;
    }
    input[type="text"]:focus, input[type="number"]:focus, input:not([type]):focus, select:focus{
      border-color:var(--ds-accent);
      box-shadow:0 0 0 4px var(--ds-accent-glow);
      background:#fff;
    }
    button{
      background:linear-gradient(135deg,#7f5cff 0%,#713DFF 52%,#5d35da 100%);
      color:#fff;
      border:1px solid rgba(113,61,255,.35);
      box-shadow:0 10px 22px rgba(113,61,255,.22);
      font-weight:700;
    }
    button:hover{
      background:linear-gradient(135deg,#8d6cff 0%,#7a49ff 54%,#643ce4 100%);
      box-shadow:0 12px 26px rgba(113,61,255,.28);
    }
    button.secondary{
      background:rgba(255,255,255,.88);
      color:#334155;
      border:1px solid rgba(0,0,0,.08);
      box-shadow:none;
      font-weight:600;
    }
    button.secondary:hover{
      background:#fff;
      border-color:rgba(0,0,0,.12);
    }
    .adv-toggle.text-link{
      color:#5b4cc4;
      font-weight:600;
      border-bottom:1px dashed rgba(91,76,196,.35);
      text-decoration:none;
    }
    .adv-toggle.text-link:hover{color:#4437a7;border-bottom-color:rgba(68,55,167,.55)}
    .link-pill{
      background:rgba(255,255,255,.88);
      border:1px solid rgba(0,0,0,.08);
      color:#334155;
      border-radius:999px;
      padding:6px 14px;
      font-weight:600;
    }
    .link-pill:hover{background:#fff;border-color:rgba(0,0,0,.13)}
    .task-table-wrap{
      background:rgba(255,255,255,.82);
      border:1px solid rgba(255,255,255,.86);
      border-radius:18px;
      overflow:hidden;
    }
    table{background:transparent;border:none}
    th{
      background:rgba(15,23,42,.04);
      color:#475569;
      font-weight:700;
      border-bottom:1px solid rgba(15,23,42,.08);
    }
    td{
      color:#1e293b;
      border-bottom:1px solid rgba(15,23,42,.06);
    }
    tbody tr:hover td{background:rgba(15,23,42,.03)}
    .material-panel,.log-panel{
      background:rgba(255,255,255,.92);
      border:1px solid rgba(255,255,255,.92);
      border-radius:20px;
      box-shadow:0 18px 48px rgba(31,38,135,.18);
      backdrop-filter:blur(18px);
      -webkit-backdrop-filter:blur(18px);
    }
    html[data-theme="dark"] body{
      background:
        radial-gradient(1200px 680px at 8% -6%, rgba(98,69,255,.26), transparent 58%),
        radial-gradient(1100px 640px at 100% 0%, rgba(64,106,198,.22), transparent 56%),
        linear-gradient(180deg,#181e2c 0%, #141b28 52%, #101624 100%);
      color:#eaf0fb;
    }
    html[data-theme="dark"] .app-shell{
      max-width:1480px;
      background:rgba(10,14,24,.18);
      border:1px solid rgba(141,166,235,.18);
      border-radius:16px;
      box-shadow:inset 0 0 0 1px rgba(146,122,255,.16), 0 0 0 1px rgba(74,116,228,.2), 0 26px 60px rgba(4,10,20,.42);
      backdrop-filter:blur(4px);
      -webkit-backdrop-filter:blur(4px);
      margin-top:2px;
    }
    html[data-theme="dark"] .card{
      background:linear-gradient(135deg,rgba(39,46,74,.86) 0%, rgba(22,31,55,.84) 52%, rgba(18,26,46,.84) 100%);
      border:1px solid rgba(136,164,236,.22);
      box-shadow:
        inset 0 1px 0 rgba(255,255,255,.08),
        inset 0 0 0 1px rgba(123,99,255,.16),
        0 20px 52px rgba(8,14,30,.34),
        0 0 36px rgba(91,76,196,.18);
    }
    html[data-theme="dark"] .hero-title{
      background:linear-gradient(135deg,#f8fbff 0%,#c7d6ef 100%);
      -webkit-background-clip:text;
      -webkit-text-fill-color:transparent;
    }
    html[data-theme="dark"] .hero-sub,
    html[data-theme="dark"] .tiny,
    html[data-theme="dark"] .hint,
    html[data-theme="dark"] .step-caption,
    html[data-theme="dark"] .task-meta{color:#b8c6de}
    html[data-theme="dark"] .section-title,
    html[data-theme="dark"] .card-title,
    html[data-theme="dark"] .label,
    html[data-theme="dark"] .channel-main{color:#edf2fb}
    html[data-theme="dark"] .step-box{
      background:rgba(44,54,82,.62) !important;
      border:1px solid rgba(174,198,245,.16) !important;
      box-shadow:inset 0 0 0 1px rgba(141,120,255,.12), 0 12px 26px rgba(7,12,24,.32);
    }
    html[data-theme="dark"] .step-box.compact{
      background:rgba(55,67,92,.48) !important;
      border:1px solid rgba(200,215,244,.14) !important;
    }
    html[data-theme="dark"] input[type="text"],
    html[data-theme="dark"] input[type="number"],
    html[data-theme="dark"] input:not([type]),
    html[data-theme="dark"] select{
      background:rgba(40,50,71,.72);
      color:#e8eefc;
      border-color:rgba(200,215,244,.2);
    }
    html[data-theme="dark"] .upload-zone{
      background:rgba(20,27,46,.52);
      border:2px dashed rgba(134,152,215,.28);
      box-shadow:inset 0 0 0 1px rgba(112,92,245,.14);
    }
    html[data-theme="dark"] .upload-zone:hover{
      border-color:rgba(138,114,255,.56);
      background:rgba(24,32,56,.62);
      box-shadow:inset 0 0 0 1px rgba(146,122,255,.24), 0 0 26px rgba(113,61,255,.2);
    }
    html[data-theme="dark"] .upload-zone .up-icon{
      color:#f2f6ff;
      background:linear-gradient(135deg,rgba(118,98,255,.22),rgba(93,124,235,.2));
      box-shadow:inset 0 0 0 1px rgba(180,203,244,.24);
    }
    html[data-theme="dark"] .upload-zone .up-title{color:#f5f8ff}
    html[data-theme="dark"] .upload-zone .up-sub{color:#97a8c8}
    html[data-theme="dark"] .theme-switch{
      background:linear-gradient(180deg,#28324a,#1f283f);
      border-color:rgba(154,177,230,.26);
      box-shadow:inset 0 1px 0 rgba(255,255,255,.08);
    }
    html[data-theme="dark"] .theme-btn{
      color:#9eb1d8;
      background:transparent;
    }
    html[data-theme="dark"] .theme-btn:hover{
      color:#f3f7ff;
      background:rgba(173,195,244,.14);
    }
    html[data-theme="dark"] .theme-btn.active{
      background:linear-gradient(135deg,#7b5eff 0%,#6b4af4 52%,#5f8dff 100%);
      color:#fff;
      box-shadow:0 10px 20px rgba(102,84,240,.32);
    }
    html[data-theme="dark"] button.secondary{
      background:rgba(60,74,103,.62);
      color:#ecf3ff;
      border-color:rgba(200,215,244,.24);
    }
    html[data-theme="dark"] .action-strip{
      background:rgba(18,25,42,.24);
      border:1px solid rgba(141,166,235,.14);
      border-radius:16px;
      padding:10px 12px;
      box-shadow:inset 0 1px 0 rgba(255,255,255,.04);
    }
    html[data-theme="dark"] .batch-op{
      background:rgba(26,35,56,.56);
      border-color:rgba(151,176,232,.2);
    }
    html[data-theme="dark"] .batch-op button.secondary{
      background:rgba(50,63,90,.72);
      border:1px solid rgba(169,190,236,.24);
      color:#edf3ff;
      box-shadow:inset 0 1px 0 rgba(255,255,255,.06);
    }
    html[data-theme="dark"] .batch-op button.secondary:hover{
      background:rgba(64,79,110,.8);
      border-color:rgba(180,203,246,.34);
    }
    html[data-theme="dark"] .batch-op .tiny{
      color:#b8c7e3;
      border-left-color:rgba(173,195,244,.24);
    }
    html[data-theme="dark"] .task-table-wrap{
      background:rgba(42,53,75,.66);
      border-color:rgba(200,215,244,.16);
    }
    html[data-theme="dark"] th{
      background:rgba(200,215,244,.08);
      color:#d6e2f5;
      border-bottom-color:rgba(200,215,244,.2);
    }
    html[data-theme="dark"] td{
      color:#eef4ff;
      border-bottom-color:rgba(200,215,244,.14);
    }
    html[data-theme="dark"] tbody tr:hover td{background:rgba(200,215,244,.09)}
    html[data-theme="dark"] .material-panel,
    html[data-theme="dark"] .log-panel{
      background:rgba(46,58,82,.88);
      border-color:rgba(200,215,244,.2);
      box-shadow:0 24px 62px rgba(8,14,28,.35);
    }
    html[data-theme="dark"] #logs{
      background:rgba(16,24,41,.76);
      border-color:rgba(157,184,240,.22);
      color:#eaf2ff;
    }

    /* ---- Light mode final pass (override legacy rules) ---- */
    html:not([data-theme="dark"]) body{
      background:
        radial-gradient(1200px 700px at 8% -8%, rgba(113,61,255,.10), transparent 58%),
        radial-gradient(1200px 700px at 100% 0%, rgba(71,135,255,.08), transparent 56%),
        linear-gradient(180deg,#f6f8fc 0%, #f3f5fa 100%);
      color:#1e293b;
    }
    html:not([data-theme="dark"]) .app-shell{
      max-width:1480px;
      background:rgba(255,255,255,.28);
      border:1px solid rgba(179,196,233,.34);
      border-radius:16px;
      box-shadow:inset 0 1px 0 rgba(255,255,255,.65), 0 18px 42px rgba(31,38,135,.08);
      backdrop-filter:blur(4px);
      -webkit-backdrop-filter:blur(4px);
      margin-top:2px;
    }
    html:not([data-theme="dark"]) .card{
      background:rgba(255,255,255,.68) !important;
      border:1px solid rgba(255,255,255,.9) !important;
      box-shadow:0 14px 36px rgba(31,38,135,.09), inset 0 1px 0 rgba(255,255,255,.7);
      border-radius:24px;
      backdrop-filter:blur(18px) saturate(116%);
      -webkit-backdrop-filter:blur(18px) saturate(116%);
    }
    html:not([data-theme="dark"]) .hero-title{
      background:linear-gradient(135deg,#1f2937 0%,#64748b 100%);
      -webkit-background-clip:text;
      -webkit-text-fill-color:transparent;
    }
    html:not([data-theme="dark"]) .hero-sub{color:#64748b}
    html:not([data-theme="dark"]) .upload-zone{
      background:rgba(255,255,255,.48);
      border:2px dashed rgba(135,120,230,.36);
      box-shadow:inset 0 0 0 1px rgba(113,61,255,.10);
    }
    html:not([data-theme="dark"]) .upload-zone:hover{
      background:rgba(255,255,255,.62);
      border-color:rgba(113,61,255,.58);
      box-shadow:inset 0 0 0 1px rgba(113,61,255,.16), 0 10px 22px rgba(113,61,255,.10);
    }
    html:not([data-theme="dark"]) .upload-zone .up-title{color:#0f172a}
    html:not([data-theme="dark"]) .upload-zone .up-sub{color:#64748b}
    html:not([data-theme="dark"]) .upload-zone .up-icon{
      color:#2f3c56;
      background:linear-gradient(135deg,rgba(113,61,255,.12),rgba(59,130,246,.12));
    }
    html:not([data-theme="dark"]) .theme-switch{
      background:linear-gradient(180deg,#ffffff,#f4f7fc);
      border-color:rgba(148,171,219,.34);
      box-shadow:inset 0 1px 0 rgba(255,255,255,.84);
    }
    html:not([data-theme="dark"]) .theme-btn{
      color:#334155;
      background:rgba(15,23,42,.05);
    }
    html:not([data-theme="dark"]) .theme-btn:hover{
      color:#111827;
      background:rgba(15,23,42,.09);
    }
    html:not([data-theme="dark"]) .theme-btn.active{
      background:linear-gradient(180deg,#ffffff,#f5f7fb);
      color:#111827;
      box-shadow:inset 0 1px 0 rgba(255,255,255,.92), 0 1px 3px rgba(15,23,42,.10);
      border:1px solid rgba(148,163,184,.34);
    }
    html:not([data-theme="dark"]) button{
      background:linear-gradient(135deg,#7f5cff 0%,#713DFF 52%,#5d35da 100%);
      border:1px solid rgba(113,61,255,.34);
      box-shadow:0 10px 22px rgba(113,61,255,.20);
      color:#fff;
      font-weight:700;
    }
    html:not([data-theme="dark"]) button:hover{
      background:linear-gradient(135deg,#8c6bff 0%,#7b4dff 54%,#6741e8 100%);
      box-shadow:0 12px 26px rgba(113,61,255,.26);
    }
    html:not([data-theme="dark"]) button.secondary{
      background:rgba(255,255,255,.90);
      color:#334155;
      border:1px solid rgba(148,163,184,.34);
      box-shadow:none;
    }
    html:not([data-theme="dark"]) button.secondary:hover{
      background:#fff;
      border-color:rgba(100,116,139,.42);
    }
    html:not([data-theme="dark"]) .action-strip{
      background:rgba(255,255,255,.54);
      border:1px solid rgba(184,199,230,.46);
      border-radius:16px;
      box-shadow:inset 0 1px 0 rgba(255,255,255,.8);
      padding:10px 12px;
    }
    html:not([data-theme="dark"]) .batch-op{
      background:rgba(255,255,255,.66);
      border:1px solid rgba(160,176,206,.46);
    }
    html:not([data-theme="dark"]) .batch-op button.secondary{
      background:linear-gradient(180deg,#ffffff,#f8fafc);
      border:1px solid rgba(148,163,184,.34);
      color:#334155;
      box-shadow:inset 0 1px 0 rgba(255,255,255,.9), 0 1px 2px rgba(15,23,42,.04);
    }
    html:not([data-theme="dark"]) .batch-op button.secondary:hover{
      background:#fff;
      border-color:rgba(100,116,139,.4);
    }
    html:not([data-theme="dark"]) .batch-op .tiny{
      color:#64748b;
      border-left-color:rgba(100,116,139,.26);
    }
    html:not([data-theme="dark"]) .task-table-wrap{
      background:rgba(255,255,255,.76);
      border:1px solid rgba(180,196,226,.46);
      border-radius:18px;
    }
    html:not([data-theme="dark"]) th{
      background:rgba(15,23,42,.04);
      color:#475569;
      border-bottom:1px solid rgba(15,23,42,.10);
      font-weight:700;
    }
    html:not([data-theme="dark"]) td{
      color:#1e293b;
      border-bottom:1px solid rgba(15,23,42,.06);
    }
    html:not([data-theme="dark"]) tbody tr:hover td{background:rgba(15,23,42,.03)}
    html:not([data-theme="dark"]) .link-pill{
      background:rgba(255,255,255,.88);
      border:1px solid rgba(100,116,139,.26);
      color:#334155;
      font-weight:600;
    }
    html:not([data-theme="dark"]) .link-pill:hover{
      background:#fff;
      border-color:rgba(71,85,105,.36);
    }
    html:not([data-theme="dark"]) .material-panel,
    html:not([data-theme="dark"]) .log-panel{
      background:rgba(255,255,255,.94);
      border:1px solid rgba(255,255,255,.9);
      box-shadow:0 22px 54px rgba(31,38,135,.16);
      backdrop-filter:blur(18px);
      -webkit-backdrop-filter:blur(18px);
    }

    /* ---- Modal system unify (log/material) ---- */
    .log-modal,.material-modal{
      padding:20px;
      backdrop-filter:blur(8px);
      -webkit-backdrop-filter:blur(8px);
    }
    .log-panel,.material-panel{
      width:min(1040px, calc(100vw - 40px));
      min-width:720px;
      max-height:88vh;
      border-radius:20px;
      padding:0;
      overflow:hidden;
      display:flex;
      flex-direction:column;
      position:relative;
    }
    .log-head{
      margin:0;
      padding:14px 16px;
      min-height:60px;
      border-bottom:1px solid rgba(100,116,139,.16);
      background:linear-gradient(180deg,rgba(255,255,255,.72),rgba(255,255,255,.52));
      flex:0 0 auto;
    }
    #logTitle,#materialTitle{
      font-size:16px;
      line-height:1.2;
      font-weight:800;
      letter-spacing:.1px;
      color:#1f2937;
    }
    #logs{
      height:auto !important;
      max-height:none !important;
      flex:1 1 auto;
      overflow:auto;
      margin:0;
      border-radius:0 0 20px 20px;
      padding:14px 16px;
      line-height:1.55;
      font-size:12px;
    }
    #materialRows{
      flex:1 1 auto;
      overflow:auto;
      padding:12px 14px 14px 14px;
    }
    .material-row{
      border-radius:14px;
      padding:12px;
      margin-bottom:10px;
    }
    .material-row{
      background:linear-gradient(180deg,rgba(255,255,255,.92),rgba(248,250,255,.86));
      border:1px solid rgba(170,184,218,.24);
      box-shadow:0 8px 20px rgba(31,38,135,.06);
    }
    .material-row h4{
      margin:0 0 10px 0;
      font-size:17px;
      line-height:1.3;
      font-weight:800;
      color:#1e293b;
      letter-spacing:.1px;
    }
    .material-grid{
      display:grid;
      grid-template-columns:1fr 1fr;
      gap:12px;
      align-items:start;
    }
    .material-upload-block{
      background:rgba(113,61,255,.06);
      border:1px solid rgba(113,61,255,.16);
      border-radius:16px;
      padding:10px 12px;
    }
    .material-upload-title{
      font-size:13px;
      font-weight:700;
      color:#475569;
      margin:0 0 8px 0;
    }
    .material-upload-block input[type="file"]{
      width:100%;
      max-width:none;
      min-width:0;
      height:42px;
      border-radius:12px;
      background:rgba(255,255,255,.86);
      border:1px solid rgba(148,163,184,.28);
      padding:6px 10px;
      color:#475569;
      line-height:28px;
    }
    .material-upload-block input[type="file"]::file-selector-button{
      height:30px;
      padding:0 14px;
      margin-right:10px;
      border:1px solid rgba(113,61,255,.26);
      border-radius:999px;
      background:linear-gradient(135deg,rgba(113,61,255,.16),rgba(59,130,246,.14));
      color:#3f2d99;
      font-size:13px;
      font-weight:700;
      cursor:pointer;
      transition:all .2s ease;
    }
    .material-upload-block input[type="file"]::file-selector-button:hover{
      background:linear-gradient(135deg,rgba(113,61,255,.22),rgba(59,130,246,.2));
      border-color:rgba(113,61,255,.34);
    }
    .material-note{
      margin:0 0 10px 0;
      padding:10px 12px;
      border-radius:12px;
      border:1px solid rgba(148,163,184,.24);
      background:linear-gradient(180deg,rgba(248,250,255,.9),rgba(243,247,255,.78));
      color:#5b6476;
      font-size:13px;
      line-height:1.6;
      box-shadow:inset 0 1px 0 rgba(255,255,255,.84);
    }
    .material-note b{
      color:#374151;
      font-weight:700;
    }
    html[data-theme="dark"] .material-row{
      background:linear-gradient(180deg,rgba(56,68,95,.72),rgba(48,59,83,.66));
      border-color:rgba(174,198,245,.18);
      box-shadow:0 10px 24px rgba(8,14,28,.26);
    }
    html[data-theme="dark"] .material-row h4{color:#eef4ff}
    html[data-theme="dark"] .material-upload-block{
      background:rgba(122,101,255,.12);
      border-color:rgba(174,198,245,.2);
    }
    html[data-theme="dark"] .material-upload-title{color:#c9d7f0}
    html[data-theme="dark"] .material-upload-block input[type="file"]{
      background:rgba(34,43,64,.72);
      border-color:rgba(174,198,245,.2);
      color:#dce6fb;
    }
    html[data-theme="dark"] .material-upload-block input[type="file"]::file-selector-button{
      background:linear-gradient(135deg,rgba(146,122,255,.3),rgba(92,137,255,.26));
      border-color:rgba(174,198,245,.28);
      color:#f1f5ff;
    }
    html[data-theme="dark"] .material-note{
      border-color:rgba(174,198,245,.2);
      background:linear-gradient(180deg,rgba(46,57,82,.7),rgba(37,48,70,.62));
      color:#b9c6df;
      box-shadow:inset 0 1px 0 rgba(255,255,255,.06);
    }
    html[data-theme="dark"] .material-note b{color:#eaf1ff}
    @media (max-width: 980px){
      .log-modal,.material-modal{padding:10px}
      .log-panel,.material-panel{
        width:calc(100vw - 20px);
        min-width:0;
        max-height:92vh;
        border-radius:14px;
      }
      .log-head{padding:10px 12px;min-height:52px}
      #logs,#materialRows{padding:10px}
      .material-grid{grid-template-columns:1fr}
    }
    html[data-theme="dark"] .log-head{
      border-bottom-color:rgba(173,195,244,.18);
      background:linear-gradient(180deg,rgba(37,47,70,.8),rgba(31,40,60,.58));
    }
    html[data-theme="dark"] #logTitle,
    html[data-theme="dark"] #materialTitle{
      color:#edf3ff;
    }
    html[data-theme="dark"] #logs{
      border-top:1px solid rgba(173,195,244,.12);
    }
    html:not([data-theme="dark"]) .log-head{
      border-bottom-color:rgba(100,116,139,.18);
      background:linear-gradient(180deg,rgba(255,255,255,.78),rgba(248,251,255,.62));
    }
  </style>
</head>
<body>
<div class="app-shell">
  <main class="main">
    <div>
      <div class="card">
        <div class="hero-bar">
          <div>
            <div class="hero-title">精准营销自动化配置（业务版）</div>
            <div class="hero-sub">选择文件 -> 添加素材 -> 开始执行</div>
          </div>
          <div class="theme-wrap">
            <div class="theme-switch" role="group" aria-label="页面显示模式">
              <button id="themeModeLight" class="theme-btn" type="button" title="浅色" data-mode="light" aria-label="浅色">
                <svg viewBox="0 0 24 24"><circle cx="12" cy="12" r="4"/><path d="M12 2.5v2.2M12 19.3v2.2M21.5 12h-2.2M4.7 12H2.5M18.9 5.1l-1.6 1.6M6.7 17.3l-1.6 1.6M18.9 18.9l-1.6-1.6M6.7 6.7 5.1 5.1"/></svg>
              </button>
              <button id="themeModeDark" class="theme-btn" type="button" title="深色" data-mode="dark" aria-label="深色">
                <svg viewBox="0 0 24 24"><path d="M20 14.3A8 8 0 1 1 9.7 4 6.5 6.5 0 1 0 20 14.3Z"/></svg>
              </button>
            </div>
          </div>
        </div>
        <div class="compose-layout">
          <div id="uploadZone" class="upload-zone" role="button" aria-label="上传任务文件">
            <div>
              <div class="up-icon">☁</div>
              <div class="up-title">点击或将文件拖拽至此</div>
              <div class="up-sub">支持 .csv, .xlsx, .zip 格式 (最大 50MB)</div>
              <input id="files" class="file-uniform hidden" type="file" multiple accept=".csv,.xlsx,.zip"/>
            </div>
          </div>
          <div id="uploadHint" class="tiny" style="margin-top:8px;min-height:20px;opacity:.9;"></div>
          <div class="config-foot">
            <div class="left">
              <button id="advToggleBtn" type="button" class="adv-toggle text-link" onclick="toggleAdvancedConfig()">高级配置（展开）</button>
            </div>
            <div class="right">
              <a class="link-pill" href="/api/template/xlsx">Excel模板</a>
              <a class="link-pill" href="/api/template/csv">CSV模板</a>
            </div>
          </div>
          <div id="advancedConfig" class="adv-panel">
            <div class="form-grid">
              <div class="field vertical">
                <label><span class="label">浏览器调试地址</span><input id="cdp_endpoint" value="http://127.0.0.1:18800" style="width:220px"/></label>
                <span class="tiny">作用：接管本地已登录浏览器（默认 127.0.0.1:18800）。</span>
              </div>
              <div class="field vertical">
                <label><span class="label">并发任务数</span><input id="concurrent" type="number" min="1" value="1" style="width:88px"/></label>
                <span class="tiny">作用：同时执行的任务数。建议先用 1 验证稳定性。</span>
              </div>
              <div class="field vertical">
                <label><span class="label">保留浏览器(秒)</span><input id="hold_seconds" type="number" min="0" value="2" style="width:88px"/></label>
                <span class="tiny">作用：任务结束后页面停留时间，便于人工复核。</span>
              </div>
              <div class="field vertical full">
                <label><span class="label">失败通知 Webhook（飞书机器人）</span><input id="notify_webhook" placeholder="https://open.feishu.cn/open-apis/bot/v2/hook/..." style="width:min(760px,95%)"/></label>
                <span class="tiny">作用：任务失败时自动推送摘要和日志尾部，便于你与地区同事协同排查。</span>
              </div>
              <div class="field vertical full">
                <label class="inline-check channel-strong"><input id="executor_include_franchise" type="checkbox" checked/> 执行员工包含加盟区域（自动同步勾选“xx加盟”节点）</label>
                <span class="tiny wrap">示例：执行员工=广佛省区，自动追加广佛省区加盟；执行员工=大郑州营运区，自动追加大郑州营运区加盟。</span>
              </div>
            </div>
          </div>
        </div>
        <div style="display:none">
          <input class="step3_channel" type="checkbox" value="短信"/>
          <input class="step3_channel" type="checkbox" value="会员通-发客户消息"/>
          <input class="step3_channel" type="checkbox" value="会员通-发送社群"/>
          <input class="step3_channel" type="checkbox" value="会员通-发客户朋友圈"/>
        </div>
      </div>

      <div class="action-strip">
        <div class="action-left">
          <button onclick="startExecute()">开始执行</button>
          <button class="secondary" onclick="retryFailed()">重试失败项</button>
        </div>
        <div class="action-right">
          <div class="batch-op">
            <button class="secondary" onclick="batchPauseSelected()">批量暂停</button>
            <button class="secondary" onclick="batchDeleteSelected()">批量删除</button>
            <span class="tiny" id="batchInfo">已选 0 项</span>
          </div>
        </div>
      </div>

      <div class="card task-section-spacer">
        <div class="task-card-head">
          <h3 class="card-title" style="margin-bottom:0">任务列表</h3>
          <div class="task-meta">按任务点击“添加素材”，补充素材后可直接执行或重试</div>
        </div>
        <div class="task-table-wrap">
          <table>
            <thead><tr>
              <th class="check-col"><input id="selectAllTasks" type="checkbox" onchange="toggleSelectAllTasks(this.checked)"/></th>
              <th>文件</th><th>计划名称</th><th>发送渠道</th><th>状态</th><th>进度</th><th>成功/失败</th><th>开始</th><th>完成</th><th>耗时(s)</th><th>操作</th>
            </tr></thead>
            <tbody id="taskRows"></tbody>
          </table>
        </div>
      </div>
    </div>
  </main>
</div>
<div id="logModal" class="log-modal" onclick="closeLogModal(event)">
  <div class="log-panel" onclick="event.stopPropagation()">
    <div class="log-head">
      <div id="logTitle">未选中任务</div>
      <button class="secondary" onclick="closeLogModal()">关闭</button>
    </div>
    <div id="logs"></div>
  </div>
</div>
<div id="materialModal" class="material-modal" onclick="closeMaterialModal(event)">
  <div class="material-panel" onclick="event.stopPropagation()">
    <div class="log-head">
      <div id="materialTitle">按任务添加素材</div>
      <div class="row">
        <button class="secondary" onclick="closeMaterialModal()">关闭</button>
        <button onclick="saveTaskMaterials()">保存修复配置</button>
      </div>
    </div>
    <div class="material-note"><b>说明：</b>可按计划补充/覆盖素材，保存后回写任务CSV；执行/重试时按最新素材运行。</div>
    <div class="row" style="margin:8px 0 6px 0;gap:10px;align-items:center;flex-wrap:wrap">
      <input id="imagePackFile" type="file" accept=".zip" class="file-uniform"/>
      <button class="secondary" onclick="uploadImagePackForTask()">批量导入图片包（按计划图片ID）</button>
      <span class="tiny">zip结构：计划图片ID/小卡.jpg（小程序封面）+ 1.jpg..9.jpg（添加图片，按序）</span>
    </div>
    <div id="materialRows"></div>
  </div>
</div>
<script>
let selectedTaskId = "";
let logOffset = 0;
let materialTaskId = "";
let materialPlans = [];
let materialTokenSeq = 0;
let uploading = false;
let selectedTaskIds = new Set();
const materialFileMap = new Map();
const LS_KEYS = {
  tasks: 'pm_ui_cached_tasks_v1',
  selectedTaskId: 'pm_ui_selected_task_id_v1',
  logsText: 'pm_ui_cached_logs_text_v1',
  logsTitle: 'pm_ui_cached_logs_title_v1',
  prefs: 'pm_ui_prefs_v1',
};

let currentThemeMode = 'light';
function applyTheme(theme){
  const root = document.documentElement;
  if(theme === 'dark') root.setAttribute('data-theme', 'dark');
  else root.removeAttribute('data-theme');
}
function updateThemeButtons(mode){
  document.querySelectorAll('.theme-btn').forEach(btn => {
    const active = btn.getAttribute('data-mode') === mode;
    btn.classList.toggle('active', active);
    btn.setAttribute('aria-pressed', active ? 'true' : 'false');
  });
}
function setThemeMode(mode, persist=true){
  const m = (mode === 'dark') ? 'dark' : 'light';
  currentThemeMode = m;
  if(persist){
    // 仅记录临时操作态；页面最终仍以时间规则自动切换
    saveLocal('pm_ui_theme_manual_last', m);
  }
  updateThemeButtons(m);
  applyTheme(m);
}
function modeByTime(){
  const now = new Date();
  const h = now.getHours();
  const m = now.getMinutes();
  const mins = h * 60 + m;
  // 07:00-17:29 浅色；17:30-次日06:59 深色
  return (mins >= 420 && mins < 1050) ? 'light' : 'dark';
}
function syncThemeByTime(){
  const mode = modeByTime();
  if(mode !== currentThemeMode){
    setThemeMode(mode, false);
  }
}

function saveLocal(key, value){
  try{ localStorage.setItem(key, JSON.stringify(value)); }catch(_){}
}
function loadLocal(key, fallback=null){
  try{
    const raw = localStorage.getItem(key);
    if(!raw) return fallback;
    return JSON.parse(raw);
  }catch(_){ return fallback; }
}

function toggleAdvancedConfig(){
  const panel = document.getElementById('advancedConfig');
  const btn = document.getElementById('advToggleBtn');
  if(!panel) return;
  panel.classList.toggle('open');
  if(btn){
    btn.textContent = panel.classList.contains('open') ? '高级配置（收起）' : '高级配置（展开）';
  }
  saveUiPrefs();
}

function selectedChannels(){
  return Array.from(document.querySelectorAll('.step3_channel:checked')).map(el => el.value);
}

function mirrorCommunityMaterialInputs(){
  const mapChecks = [
    ['msg_add_mini_program', 'msg_add_mini_program_community'],
    ['moments_add_images', 'moments_add_images_community'],
  ];
  for(const [mainId, communityId] of mapChecks){
    const main = document.getElementById(mainId);
    const community = document.getElementById(communityId);
    if(!main || !community) continue;
    if(community.checked !== main.checked) community.checked = main.checked;
  }
}

function syncChannelMaterials(){
  const channels = selectedChannels();
  const isCommunity = channels.includes('会员通-发送社群');
  const showMoments = channels.includes('会员通-发客户朋友圈') || isCommunity;
  const showMsg = channels.includes('会员通-发客户消息') || isCommunity;
  const momentsBox = document.getElementById('materialMoments');
  const msgBox = document.getElementById('materialMiniProgram');
  const momentsCommunityBox = document.getElementById('materialMomentsCommunity');
  const msgCommunityBox = document.getElementById('materialMiniProgramCommunity');
  const emptyTip = document.getElementById('materialEmptyTip');
  if(momentsBox) momentsBox.classList.toggle('hidden', !showMoments || isCommunity);
  if(msgBox) msgBox.classList.toggle('hidden', !showMsg || isCommunity);
  if(momentsCommunityBox) momentsCommunityBox.classList.toggle('hidden', !isCommunity);
  if(msgCommunityBox) msgCommunityBox.classList.toggle('hidden', !isCommunity);

  mirrorCommunityMaterialInputs();
  if(!showMoments){
    const chk = document.getElementById('moments_add_images');
    const file = document.getElementById('moments_images');
    const chkCommunity = document.getElementById('moments_add_images_community');
    const fileCommunity = document.getElementById('moments_images_community');
    if(chk) chk.checked = false;
    if(file) file.value = '';
    if(chkCommunity) chkCommunity.checked = false;
    if(fileCommunity) fileCommunity.value = '';
  }
  if(!showMsg){
    const chk = document.getElementById('msg_add_mini_program');
    const cover = document.getElementById('mini_program_cover');
    const chkCommunity = document.getElementById('msg_add_mini_program_community');
    const coverCommunity = document.getElementById('mini_program_cover_community');
    if(chk) chk.checked = false;
    if(cover) cover.value = '';
    if(chkCommunity) chkCommunity.checked = false;
    if(coverCommunity) coverCommunity.value = '';
  }
  if(emptyTip) emptyTip.style.display = (showMoments || showMsg) ? 'none' : 'block';
  saveUiPrefs();
}

function esc(s){
  return (s||"").replace(/[&<>"']/g, m => ({
    "&":"&amp;",
    "<":"&lt;",
    ">":"&gt;",
    '"':"&quot;",
    "'":"&#39;"
  }[m]));
}

function setUploadHint(msg, isOk=true){
  const el = document.getElementById('uploadHint');
  if(!el) return;
  el.textContent = msg || '';
  if(!msg){
    el.style.color = '';
    return;
  }
  el.style.color = isOk ? '#22c55e' : '#ef4444';
}

async function upload(){
  if(uploading) return;
  const fileInput = document.getElementById('files');
  const files = fileInput.files;
  if(!files.length){
    // 更顺手：未选文件时直接拉起文件选择框，而不是报错中断
    fileInput.click();
    return;
  }
  setUploadHint(`正在上传 ${files.length} 个文件并生成任务...`, true);
  uploading = true;
  fileInput.disabled = true;
  const fd = new FormData();
  for(const f of files){ fd.append('files', f); }
  // 系统默认复用已登录浏览器（固定开启）
  fd.append('connect_cdp', 'true');
  fd.append('cdp_endpoint', document.getElementById('cdp_endpoint').value);
  // 严格第2步由文件/渠道预设逻辑处理（前端不再暴露开关）
  fd.append('strict_step2', 'true');
  fd.append('skip_step2', 'false');
  fd.append('concurrent', document.getElementById('concurrent').value || '1');
  fd.append('start', '');
  fd.append('end', '');
  fd.append('hold_seconds', document.getElementById('hold_seconds').value || '2');
  fd.append('notify_webhook', document.getElementById('notify_webhook')?.value || '');
  // 渠道统一以任务文件每行“发送渠道”为准，避免全局覆盖导致跨渠道必填串扰
  fd.append('step3_channels', '');
  fd.append('executor_include_franchise', document.getElementById('executor_include_franchise').checked ? 'true' : 'false');
  // 素材已迁移至任务列表“添加素材”，这里默认不携带全局素材
  fd.append('moments_add_images', 'false');
  const momentImgs = [];
  for(const img of momentImgs){ fd.append('moments_images', img); }
  // 上传门店是否生效由任务文件字段控制，这里不做全局强制
  fd.append('upload_stores', 'false');
  fd.append('msg_add_mini_program', 'false');
  saveUiPrefs();
  try{
    const r = await fetch('/api/tasks/upload', {method:'POST', body:fd});
    if(!r.ok){
      const errText = await r.text();
      setUploadHint(`上传失败：${errText}`, false);
      alert(errText);
    } else {
      const data = await r.json();
      await refreshTasks();
      const zipApplied = Array.isArray(data.zip_applied) ? data.zip_applied.length : 0;
      if(zipApplied > 0){
        setUploadHint(`上传成功，已生成任务并自动写入图片包素材（${zipApplied} 条任务应用）。`, true);
      }else{
        setUploadHint('上传文件成功，已加入任务列表。', true);
      }
    }
  } finally {
    fileInput.value = '';
    fileInput.disabled = false;
    uploading = false;
  }
}

async function startExecute(){
  const r = await fetch('/api/tasks/start', {method:'POST'});
  if(!r.ok){ alert(await r.text()); return; }
  const data = await r.json();
  if(!data.count){
    alert('当前没有待执行任务。');
    return;
  }
  alert(`已开始执行 ${data.count} 个任务。`);
  await refreshTasks();
}

async function retryTask(id){
  await fetch('/api/tasks/' + id + '/retry', {method:'POST'});
  await refreshTasks();
}

async function pauseTask(id){
  const r = await fetch('/api/tasks/' + id + '/pause', {method:'POST'});
  if(!r.ok){ alert(await r.text()); return; }
  await refreshTasks();
}

async function resumeTask(id){
  const r = await fetch('/api/tasks/' + id + '/resume', {method:'POST'});
  if(!r.ok){ alert(await r.text()); return; }
  await refreshTasks();
}

async function deleteTask(id){
  if(!confirm('确认删除该任务？删除后不会执行。')) return;
  const r = await fetch('/api/tasks/' + id + '/delete', {method:'POST'});
  if(!r.ok){ alert(await r.text()); return; }
  selectedTaskIds.delete(id);
  await refreshTasks();
}

async function retryFailed(){
  await fetch('/api/tasks/retry-failed', {method:'POST'});
  await refreshTasks();
}

async function batchPauseSelected(){
  const ids = Array.from(selectedTaskIds);
  if(!ids.length){ alert('请先勾选任务'); return; }
  const r = await fetch('/api/tasks/pause-batch', {
    method:'POST',
    headers:{'Content-Type':'application/json'},
    body: JSON.stringify(ids),
  });
  if(!r.ok){ alert(await r.text()); return; }
  const data = await r.json();
  alert(`批量暂停完成：成功 ${data.changed || 0}，忽略 ${data.ignored || 0}`);
  await refreshTasks();
}

async function batchDeleteSelected(){
  const ids = Array.from(selectedTaskIds);
  if(!ids.length){ alert('请先勾选任务'); return; }
  if(!confirm(`确认删除已选 ${ids.length} 个任务？删除后不会执行。`)) return;
  const r = await fetch('/api/tasks/delete-batch', {
    method:'POST',
    headers:{'Content-Type':'application/json'},
    body: JSON.stringify(ids),
  });
  if(!r.ok){ alert(await r.text()); return; }
  const data = await r.json();
  alert(`批量删除完成：成功 ${data.changed || 0}，忽略 ${data.ignored || 0}`);
  selectedTaskIds.clear();
  await refreshTasks();
}

function fmtStatus(s){return '<span class="status-'+s+'">'+s+'</span>';}

function renderFileLink(task){
  return `<a class="link-pill" href="/api/tasks/${task.id}/file">下载CSV</a>`;
}

function updateBatchInfo(){
  const n = selectedTaskIds.size;
  const el = document.getElementById('batchInfo');
  if(el) el.textContent = `已选 ${n} 项`;
  const all = document.getElementById('selectAllTasks');
  if(all){
    const boxes = Array.from(document.querySelectorAll('.task-select'));
    all.checked = boxes.length > 0 && boxes.every(b => b.checked);
  }
}

function toggleTaskSelection(id, checked){
  if(checked) selectedTaskIds.add(id);
  else selectedTaskIds.delete(id);
  updateBatchInfo();
}

function toggleSelectAllTasks(checked){
  document.querySelectorAll('.task-select').forEach(el => {
    el.checked = checked;
    const id = el.getAttribute('data-id');
    if(!id) return;
    if(checked) selectedTaskIds.add(id);
    else selectedTaskIds.delete(id);
  });
  updateBatchInfo();
}

function _statusText(t){
  if(t.paused && t.status === 'pending') return 'paused';
  return t.status;
}

function renderTasks(rows){
  const visibleIds = new Set(rows.map(r => r.id));
  selectedTaskIds = new Set(Array.from(selectedTaskIds).filter(id => visibleIds.has(id)));
  const tbody = document.getElementById('taskRows');
  tbody.innerHTML = rows.map(t => `
    <tr>
      <td class="check-col"><input class="task-select" data-id="${t.id}" type="checkbox" ${selectedTaskIds.has(t.id) ? 'checked' : ''} onchange="toggleTaskSelection('${t.id}', this.checked)"/></td>
      <td><div class="file-hero">${esc(t.filename)}</div><div style="margin-top:4px">${renderFileLink(t)}</div></td>
      <td>${esc(t.plan_name || '-')}</td>
      <td>${esc(t.send_channels || '-')}</td>
      <td>${fmtStatus(_statusText(t))}</td>
      <td>${t.completed_plans}/${t.total_plans || '-'}</td>
      <td>${t.success_count}/${t.fail_count}</td>
      <td>${esc(t.started_at || '-')}</td>
      <td>${esc(t.ended_at || '-')}</td>
      <td>${t.duration_sec ? t.duration_sec.toFixed(1) : '-'}</td>
      <td>
        <button onclick="openLogModal('${t.id}')">日志</button>
        <button class="secondary" onclick="openMaterialModal('${t.id}')">添加素材</button>
        ${t.status === 'running' ? '' : (t.paused && t.status === 'pending' ? `<button class="secondary" onclick="resumeTask('${t.id}')">恢复</button>` : `<button class="secondary" onclick="pauseTask('${t.id}')">暂停</button>`)}
        ${t.status === 'running' ? '' : `<button class="secondary" onclick="deleteTask('${t.id}')">删除</button>`}
        ${t.status === 'failed' ? `<button class="secondary" onclick="retryTask('${t.id}')">重试</button>` : ''}
      </td>
    </tr>
  `).join('');
  updateBatchInfo();
}

async function refreshTasks(){
  let rows = [];
  try{
    const r = await fetch('/api/tasks');
    const data = await r.json();
    rows = data.tasks || [];
    saveLocal(LS_KEYS.tasks, rows);
  }catch(_){
    rows = loadLocal(LS_KEYS.tasks, []);
  }
  if(!selectedTaskId){
    const running = rows.find(t => t.status === 'running');
    if(running){
      selectedTaskId = running.id;
      saveLocal(LS_KEYS.selectedTaskId, selectedTaskId);
      logOffset = 0;
      document.getElementById('logs').textContent = "";
      document.getElementById('logTitle').textContent = `任务 ${running.filename} (${running.status})`;
      saveLocal(LS_KEYS.logsText, "");
      saveLocal(LS_KEYS.logsTitle, document.getElementById('logTitle').textContent);
    }
  }else{
    const exists = rows.some(t => t.id === selectedTaskId);
    if(!exists){
      selectedTaskId = "";
      logOffset = 0;
      saveLocal(LS_KEYS.selectedTaskId, "");
      document.getElementById('logTitle').textContent = '任务不存在（可能服务重启），请重新选择';
    }
  }
  renderTasks(rows);
}

function openLogModal(id){
  const m = document.getElementById('logModal');
  if(m) m.classList.add('open');
  return selectTask(id);
}

function closeLogModal(evt){
  if(evt && evt.target && evt.target.id !== 'logModal') return;
  const m = document.getElementById('logModal');
  if(m) m.classList.remove('open');
}

function _channelList(raw){
  return String(raw || '').split(/[|,，、/]/).map(s => s.trim()).filter(Boolean);
}

function _supportMini(channels){
  const s = new Set(_channelList(channels));
  return s.has('会员通-发客户消息') || s.has('会员通-发送社群');
}

function _supportMoments(channels){
  const s = new Set(_channelList(channels));
  return s.has('会员通-发客户朋友圈') || s.has('会员通-发送社群');
}

function closeMaterialModal(evt){
  if(evt && evt.target && evt.target.id !== 'materialModal') return;
  const m = document.getElementById('materialModal');
  if(m) m.classList.remove('open');
}

function _renderMaterialPreviews(row, idx){
  const box = row.querySelector(`.img-preview[data-kind="moments"][data-idx="${idx}"]`);
  if(!box) return;
  box.innerHTML = '';
  (materialPlans[idx]?.moment_tokens || []).forEach(tok => {
    const f = materialFileMap.get(tok);
    if(!f) return;
    const img = document.createElement('img');
    img.src = URL.createObjectURL(f);
    img.alt = f.name;
    box.appendChild(img);
  });
}

function renderMaterialRows(){
  const root = document.getElementById('materialRows');
  if(!root) return;
  if(!materialPlans.length){
    root.innerHTML = '<div class="tiny">当前任务没有可配置计划行。</div>';
    return;
  }
  root.innerHTML = materialPlans.map(p => {
    const mini = _supportMini(p.channels);
    const moments = _supportMoments(p.channels);
    const idx = p.index;
    const chips = (p.moments_image_paths || '').split('|').filter(Boolean).map(x => `<span class="path-chip">${esc(x.split('/').pop())}</span>`).join('');
    return `
      <div class="material-row" data-idx="${idx}">
        <h4>计划${idx + 1}：${esc(p.name)} <span class="tiny">（渠道：${esc(p.channels || '-')}，计划图片ID：${esc(p.plan_image_id || '-')}）</span></h4>
        <div class="material-grid">
          <div class="material-upload-block ${mini ? '' : 'hidden'}">
            <div class="material-upload-title">小程序封面</div>
            <div>
              <input type="file" data-kind="mini-cover" data-idx="${idx}" accept=".jpg,.jpeg,.png"/>
              ${p.msg_mini_program_cover_path ? `<div class="tiny" style="margin-top:4px">当前封面：${esc(p.msg_mini_program_cover_path.split('/').pop())}</div>` : ''}
            </div>
          </div>
          <div class="material-upload-block ${moments ? '' : 'hidden'}">
            <div class="material-upload-title">朋友圈图片</div>
            <div>
              <input type="file" data-kind="moments-files" data-idx="${idx}" multiple accept=".jpg,.jpeg,.png"/>
              <div class="img-preview" data-kind="moments" data-idx="${idx}"></div>
              ${chips ? `<div style="margin-top:4px">${chips}</div>` : ''}
            </div>
          </div>
        </div>
      </div>
    `;
  }).join('');

  root.querySelectorAll('input[data-kind="mini-cover"]').forEach(el => {
    el.addEventListener('change', (e) => {
      const idx = Number(e.target.dataset.idx);
      const f = e.target.files && e.target.files[0];
      if(!f) return;
      const token = `cover_${Date.now()}_${++materialTokenSeq}_${idx}`;
      materialFileMap.set(token, f);
      materialPlans[idx].cover_token = token;
    });
  });
  root.querySelectorAll('input[data-kind="moments-files"]').forEach(el => {
    el.addEventListener('change', (e) => {
      const idx = Number(e.target.dataset.idx);
      const list = Array.from(e.target.files || []);
      const tokens = [];
      list.forEach((f, i) => {
        const token = `mom_${Date.now()}_${++materialTokenSeq}_${idx}_${i}`;
        materialFileMap.set(token, f);
        tokens.push(token);
      });
      materialPlans[idx].moment_tokens = tokens;
      _renderMaterialPreviews(el.closest('.material-row'), idx);
    });
  });
}

async function openMaterialModal(id){
  materialTaskId = id;
  materialFileMap.clear();
  const resp = await fetch('/api/tasks/' + id + '/plans');
  if(!resp.ok){ alert('获取计划列表失败'); return; }
  const data = await resp.json();
  materialPlans = (data.plans || []).map(p => ({
    index: Number(p.index),
    name: p.name || '',
    channels: p.channels || '',
    plan_image_id: p.plan_image_id || '',
    msg_add_mini_program: !!p.msg_add_mini_program,
    moments_add_images: !!p.moments_add_images,
    msg_mini_program_cover_path: p.msg_mini_program_cover_path || '',
    moments_image_paths: p.moments_image_paths || '',
    cover_token: '',
    moment_tokens: []
  }));
  document.getElementById('materialTitle').textContent = `按任务添加素材（任务 ${id.slice(0,8)}）`;
  renderMaterialRows();
  const m = document.getElementById('materialModal');
  if(m) m.classList.add('open');
}

async function uploadImagePackForTask(){
  if(!materialTaskId){
    alert('请先选择任务');
    return;
  }
  const inp = document.getElementById('imagePackFile');
  const f = inp && inp.files && inp.files[0];
  if(!f){
    alert('请先选择zip图片包');
    return;
  }
  const fd = new FormData();
  fd.append('image_pack', f, f.name || 'images.zip');
  const resp = await fetch(`/api/tasks/${materialTaskId}/materials/image-pack`, {method:'POST', body:fd});
  if(!resp.ok){
    alert(await resp.text());
    return;
  }
  const data = await resp.json();
  alert(`图片包导入完成：匹配 ${data.matched_plans || 0} 条，更新 ${data.updated_plans || 0} 条（小程序 ${data.updated_mini || 0}，图片 ${data.updated_moments || 0}）`);
  await openMaterialModal(materialTaskId);
}

async function saveTaskMaterials(){
  if(!materialTaskId){ return; }
  const specs = materialPlans.map(p => ({
    index: p.index,
    msg_add_mini_program: !!p.msg_add_mini_program,
    moments_add_images: !!p.moments_add_images,
    cover_token: p.cover_token || '',
    moment_tokens: p.moment_tokens || [],
  }));
  const fd = new FormData();
  fd.append('specs_json', JSON.stringify(specs));
  materialFileMap.forEach((file, token) => {
    fd.append('files', file, token);
  });
  const resp = await fetch(`/api/tasks/${materialTaskId}/materials`, {method:'POST', body:fd});
  if(!resp.ok){
    alert(await resp.text());
    return;
  }
  alert('素材已保存，可直接重试该任务执行。');
  closeMaterialModal();
}

async function selectTask(id){
  selectedTaskId = id;
  saveLocal(LS_KEYS.selectedTaskId, selectedTaskId);
  logOffset = 0;
  document.getElementById('logs').textContent = "";
  const resp = await fetch('/api/tasks/' + id);
  if(resp.status === 404){
    selectedTaskId = "";
    saveLocal(LS_KEYS.selectedTaskId, "");
    document.getElementById('logTitle').textContent = '任务不存在（可能服务重启），请重新选择';
    return;
  }
  const t = await resp.json();
  document.getElementById('logTitle').textContent = `任务 ${t.filename} (${t.status})`;
  saveLocal(LS_KEYS.logsTitle, document.getElementById('logTitle').textContent);
  saveLocal(LS_KEYS.logsText, "");
  await pollLogs(true);
}

async function pollLogs(reset=false){
  if(!selectedTaskId) return;
  const r = await fetch(`/api/tasks/${selectedTaskId}/logs?offset=${logOffset}&limit=500`);
  if(r.status === 404){
    selectedTaskId = "";
    logOffset = 0;
    saveLocal(LS_KEYS.selectedTaskId, "");
    document.getElementById('logTitle').textContent = '任务不存在（可能服务重启），已清除旧任务选择';
    return;
  }
  const data = await r.json();
  const logs = data.logs || [];
  if(logs.length){
    const box = document.getElementById('logs');
    box.textContent += logs.join("\\n") + "\\n";
    box.scrollTop = box.scrollHeight;
    logOffset = data.next_offset || (logOffset + logs.length);
    saveLocal(LS_KEYS.logsText, box.textContent);
    saveLocal(LS_KEYS.logsTitle, document.getElementById('logTitle').textContent);
  }
}

function saveUiPrefs(){
  const momentsChecked = !!document.getElementById('moments_add_images')?.checked || !!document.getElementById('moments_add_images_community')?.checked;
  const miniProgramChecked = !!document.getElementById('msg_add_mini_program')?.checked || !!document.getElementById('msg_add_mini_program_community')?.checked;
  const prefs = {
    cdp_endpoint: document.getElementById('cdp_endpoint')?.value || '',
    concurrent: document.getElementById('concurrent')?.value || '1',
    hold_seconds: document.getElementById('hold_seconds')?.value || '2',
    notify_webhook: document.getElementById('notify_webhook')?.value || '',
    channels: selectedChannels(),
    executor_include_franchise: !!document.getElementById('executor_include_franchise')?.checked,
    moments_add_images: momentsChecked,
    msg_add_mini_program: miniProgramChecked,
    advanced_open: document.getElementById('advancedConfig')?.classList.contains('open') || false,
  };
  saveLocal(LS_KEYS.prefs, prefs);
}

function restoreUiFromCache(){
  const prefs = loadLocal(LS_KEYS.prefs, null);
  if(prefs){
    if(document.getElementById('cdp_endpoint')) document.getElementById('cdp_endpoint').value = prefs.cdp_endpoint || 'http://127.0.0.1:18800';
    if(document.getElementById('concurrent')) document.getElementById('concurrent').value = prefs.concurrent || '1';
    if(document.getElementById('hold_seconds')) document.getElementById('hold_seconds').value = prefs.hold_seconds || '2';
    if(document.getElementById('notify_webhook')) document.getElementById('notify_webhook').value = prefs.notify_webhook || '';
    const channels = new Set(prefs.channels || []);
    document.querySelectorAll('.step3_channel').forEach(el => { el.checked = channels.has(el.value); });
    if(document.getElementById('executor_include_franchise')) document.getElementById('executor_include_franchise').checked = !!prefs.executor_include_franchise;
    if(document.getElementById('moments_add_images')) document.getElementById('moments_add_images').checked = !!prefs.moments_add_images;
    if(document.getElementById('msg_add_mini_program')) document.getElementById('msg_add_mini_program').checked = !!prefs.msg_add_mini_program;
    if(prefs.advanced_open){
      const panel = document.getElementById('advancedConfig');
      const btn = document.getElementById('advToggleBtn');
      if(panel) panel.classList.add('open');
      if(btn) btn.textContent = '高级配置（收起）';
    }
  }
  setThemeMode('light', false);
  const cachedRows = loadLocal(LS_KEYS.tasks, []);
  if(cachedRows.length){
    renderTasks(cachedRows);
  }
  const cachedSelected = loadLocal(LS_KEYS.selectedTaskId, '');
  if(cachedSelected){
    selectedTaskId = cachedSelected;
  }
  const cachedTitle = loadLocal(LS_KEYS.logsTitle, '未选中任务');
  const cachedLogs = loadLocal(LS_KEYS.logsText, '');
  document.getElementById('logTitle').textContent = cachedTitle;
  document.getElementById('logs').textContent = cachedLogs;
}

setInterval(async ()=>{ await refreshTasks(); await pollLogs(); }, 2000);
document.querySelectorAll('.step3_channel').forEach(el => el.addEventListener('change', syncChannelMaterials));
const filesEl = document.getElementById('files');
if(filesEl){
  filesEl.addEventListener('change', () => {
    if(filesEl.files && filesEl.files.length){
      upload();
    }
  });
}
const uploadZoneEl = document.getElementById('uploadZone');
if(uploadZoneEl && filesEl){
  uploadZoneEl.addEventListener('click', () => filesEl.click());
  ['dragenter','dragover'].forEach(evt => {
    uploadZoneEl.addEventListener(evt, (e) => {
      e.preventDefault();
      e.stopPropagation();
      uploadZoneEl.classList.add('drag-over');
    });
  });
  ['dragleave','drop'].forEach(evt => {
    uploadZoneEl.addEventListener(evt, (e) => {
      e.preventDefault();
      e.stopPropagation();
      if(evt === 'drop'){
        const dt = e.dataTransfer;
        if(dt && dt.files && dt.files.length){
          filesEl.files = dt.files;
          upload();
        }
      }
      uploadZoneEl.classList.remove('drag-over');
    });
  });
}
['cdp_endpoint','concurrent','hold_seconds','notify_webhook','executor_include_franchise','moments_add_images','msg_add_mini_program']
  .forEach(id => {
    const el = document.getElementById(id);
    if(el){ el.addEventListener('change', saveUiPrefs); el.addEventListener('input', saveUiPrefs); }
  });
document.querySelectorAll('.theme-btn').forEach(btn => {
  btn.addEventListener('click', () => {
    const mode = btn.getAttribute('data-mode') || 'light';
    setThemeMode(mode, true);
  });
});
[['moments_add_images_community','moments_add_images'], ['msg_add_mini_program_community','msg_add_mini_program']].forEach(([fromId, toId]) => {
  const from = document.getElementById(fromId);
  const to = document.getElementById(toId);
  if(from && to){
    from.addEventListener('change', () => { to.checked = !!from.checked; saveUiPrefs(); });
    to.addEventListener('change', () => { from.checked = !!to.checked; saveUiPrefs(); });
  }
});
['moments_images_community','mini_program_cover_community'].forEach(id => {
  const el = document.getElementById(id);
  if(el){ el.addEventListener('change', saveUiPrefs); }
});
restoreUiFromCache();
syncChannelMaterials();
refreshTasks();
</script>
</body>
</html>
"""
